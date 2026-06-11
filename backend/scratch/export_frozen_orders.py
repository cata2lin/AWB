"""
Export all Frisbo orders with a FROZEN status to an .xlsx.

"Frozen" = the Frisbo aggregated_status is NON-terminal (not delivered / returned /
cancelled per the canonical classifier) AND the order is older than a threshold
(default 14 days), so it has had ample time to reach a final state in normal COD
flow but Frisbo never advanced it. Frisbo is AWB's only status source, so these are
the orders most likely showing a stale status the courier has already settled.

Run: cd backend && ./venv/Scripts/python.exe -u scratch/export_frozen_orders.py [days] [out.xlsx]
Read-only on the DB.
"""

import os
import sys
import asyncio
from datetime import datetime, timezone

sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, r"c:/Users/Admin/Desktop/AWB Print/awb-print-manager/backend")

from sqlalchemy import text  # noqa: E402
from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Font, PatternFill, Alignment  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

from app.core.database import AsyncSessionLocal  # noqa: E402
from app.core.status_classification import (  # noqa: E402
    classify,
    DELIVERED,
    RETURNED,
    REFUSED,
    CANCELLED,
)
from app.core.timezone import to_bucharest_date  # noqa: E402

DAYS = int(sys.argv[1]) if len(sys.argv) > 1 else 14
OUT = (
    sys.argv[2]
    if len(sys.argv) > 2
    else r"c:/Users/Admin/Desktop/frozen_frisbo_orders.xlsx"
)

TERMINAL = DELIVERED | RETURNED | REFUSED | CANCELLED

HEADERS = [
    "Order #",
    "Store",
    "Frisbo Status",
    "Category",
    "Freeze Type",
    "Created (RO)",
    "Days Old",
    "Has AWB",
    "Total",
    "Currency",
    "Note",
]


async def fetch():
    # Pull every non-terminal order older than the threshold.
    sql = text(
        """
        SELECT o.order_number, COALESCE(s.name, o.store_uid) AS store,
               o.aggregated_status, o.tracking_number, o.frisbo_created_at,
               o.total_price, o.currency, o.note
        FROM orders o
        LEFT JOIN stores s ON s.uid = o.store_uid
        WHERE o.frisbo_created_at IS NOT NULL
          AND o.frisbo_created_at < (now() at time zone 'utc') - make_interval(days => :d)
          AND lower(coalesce(o.aggregated_status, '')) <> ALL(:term)
        ORDER BY o.frisbo_created_at ASC
        """
    )
    async with AsyncSessionLocal() as db:
        rows = (
            await db.execute(sql, {"d": DAYS, "term": [t.lower() for t in TERMINAL]})
        ).all()
    return rows


def build_rows(rows):
    now = datetime.now(timezone.utc).replace(tzinfo=None)
    out = []
    for r in rows:
        cat = classify(r.aggregated_status)
        if cat in ("delivered", "returned", "cancelled"):
            continue  # safety net — terminal already excluded in SQL
        days_old = (now - r.frisbo_created_at).days
        freeze = (
            "Frozen in transit"
            if cat == "in_transit"
            else "Frozen pre-delivery (not shipped)"
        )
        out.append(
            [
                r.order_number,
                r.store,
                r.aggregated_status,
                cat,
                freeze,
                str(to_bucharest_date(r.frisbo_created_at) or ""),
                days_old,
                "Yes" if (r.tracking_number or "").strip() else "No",
                round(float(r.total_price or 0), 2),
                r.currency or "RON",
                (r.note or "")[:200],
            ]
        )
    # Most-frozen first
    out.sort(key=lambda x: x[6], reverse=True)
    return out


def write_xlsx(data):
    wb = Workbook()

    # --- Sheet 1: the orders ---
    ws = wb.active
    ws.title = "Frozen Orders"
    head_fill = PatternFill("solid", fgColor="1F2937")
    head_font = Font(bold=True, color="FFFFFF")
    ws.append(HEADERS)
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal="center")
    transit_fill = PatternFill("solid", fgColor="FEF3C7")  # amber = in transit
    pre_fill = PatternFill("solid", fgColor="E0E7FF")  # indigo = pre-delivery
    for row in data:
        ws.append(row)
        rfill = transit_fill if row[3] == "in_transit" else pre_fill
        ws.cell(row=ws.max_row, column=4).fill = rfill
        ws.cell(row=ws.max_row, column=5).fill = rfill
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"
    widths = [16, 18, 30, 12, 30, 13, 9, 8, 10, 9, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # --- Sheet 2: summary ---
    s = wb.create_sheet("Summary")
    s.append(["FROZEN FRISBO ORDERS — orders stuck in a non-terminal status"])
    s["A1"].font = Font(bold=True, size=13)
    s.append([f"Criterion: non-terminal Frisbo status AND older than {DAYS} days"])
    s.append(
        [f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"]
    )
    s.append([f"Total frozen orders: {len(data)}"])
    s.append([])

    from collections import Counter

    s.append(["By Frisbo status", "Count"])
    s.cell(s.max_row, 1).font = Font(bold=True)
    s.cell(s.max_row, 2).font = Font(bold=True)
    for st, n in Counter(r[2] for r in data).most_common():
        s.append([st, n])
    s.append([])
    s.append(["By freeze type", "Count"])
    s.cell(s.max_row, 1).font = Font(bold=True)
    s.cell(s.max_row, 2).font = Font(bold=True)
    for ft, n in Counter(r[4] for r in data).most_common():
        s.append([ft, n])
    s.append([])
    s.append(["By store", "Count"])
    s.cell(s.max_row, 1).font = Font(bold=True)
    s.cell(s.max_row, 2).font = Font(bold=True)
    for store, n in Counter(r[1] for r in data).most_common():
        s.append([store, n])
    s.column_dimensions["A"].width = 40
    s.column_dimensions["B"].width = 12

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)


async def main():
    rows = await fetch()
    data = build_rows(rows)
    write_xlsx(data)
    print(
        f"Frozen orders: {len(data):,}  (non-terminal status, older than {DAYS} days)"
    )
    print(f"WROTE {OUT}")


if __name__ == "__main__":
    asyncio.run(main())
