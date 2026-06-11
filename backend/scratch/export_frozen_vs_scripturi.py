"""
Export Frisbo orders with a FROZEN status, CROSS-CHECKED against Scripturi's
courier-resolved status — so each frozen order shows whether the courier actually
settled it (i.e. Frisbo's status is genuinely stale) or whether Scripturi also has
it unresolved (never shipped / still in transit / untracked).

"Frozen" = non-terminal Frisbo aggregated_status (not delivered/returned/cancelled)
AND older than `days` (default 14). Scripturi status comes from its profit_orders
(snapshot, joined on order_name == AWB order_number).

Run: cd backend && ./venv/Scripts/python.exe -u scratch/export_frozen_vs_scripturi.py [days] [out.xlsx] [scripturi.db]
Read-only on both sides.
"""

import os
import sys
import asyncio
import sqlite3
from collections import Counter
from datetime import datetime, timezone

sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, r"c:/Users/Admin/Desktop/AWB Print/awb-print-manager/backend")

from sqlalchemy import text  # noqa: E402
from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Font, PatternFill, Alignment  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

from app.core.database import AsyncSessionLocal  # noqa: E402
from app.core.status_classification import (  # noqa: E402
    classify,
    DELIVERED,
    RETURNED,
    REFUSED,
    CANCELLED,
)
from app.core.timezone import to_bucharest_date  # noqa: E402

DAYS = int(sys.argv[1]) if len(sys.argv) > 1 else 14
OUT = (
    sys.argv[2]
    if len(sys.argv) > 2
    else r"c:/Users/Admin/Desktop/frozen_frisbo_orders_vs_scripturi.xlsx"
)
SC_DB = (
    sys.argv[3]
    if len(sys.argv) > 3
    else r"c:/tmp/scr_new2/Scripturi/data/profitability.db"
)

TERMINAL = DELIVERED | RETURNED | REFUSED | CANCELLED

# Scripturi status_category -> (verdict, sort-priority). Lower priority sorts first.
SC_RESOLVED = {
    "Livrata": "Delivered",
    "Anulata": "Cancelled",
    "Refuzata": "Returned/Refused",
}


def verdict(sc_status):
    if sc_status is None:
        return ("Not in Scripturi (no courier data)", 3)
    if sc_status in SC_RESOLVED:
        return (f"FRISBO STALE — courier: {SC_RESOLVED[sc_status]}", 0)
    if sc_status == "In curs de livrare":
        return ("Both still in transit", 2)
    if sc_status == "Netrimisa":
        return ("Both: never shipped", 2)
    if sc_status == "Lipsa awb":
        return ("Scripturi has no AWB (untracked)", 2)
    return (f"Other: {sc_status}", 2)


HEADERS = [
    "Order #",
    "Store",
    "Frisbo Status",
    "Category",
    "Created (RO)",
    "Days Old",
    "Has AWB",
    "Total",
    "Currency",
    "Scripturi Status",
    "Verdict",
    "Note",
]


def load_scripturi():
    con = sqlite3.connect(SC_DB)
    con.row_factory = sqlite3.Row
    sc = {}
    for r in con.execute("SELECT order_name, status_category FROM profit_orders"):
        sc[r["order_name"]] = r["status_category"]
    con.close()
    return sc


async def fetch_frozen():
    sql = text(
        """
        SELECT o.order_number, COALESCE(s.name, o.store_uid) AS store,
               o.aggregated_status, o.tracking_number, o.frisbo_created_at,
               o.total_price, o.currency, o.note
        FROM orders o
        LEFT JOIN stores s ON s.uid = o.store_uid
        WHERE o.frisbo_created_at IS NOT NULL
          AND o.frisbo_created_at < (now() at time zone 'utc') - make_interval(days => :d)
          AND lower(coalesce(o.aggregated_status, '')) <> ALL(:term)
        ORDER BY o.frisbo_created_at ASC
        """
    )
    async with AsyncSessionLocal() as db:
        return (
            await db.execute(sql, {"d": DAYS, "term": [t.lower() for t in TERMINAL]})
        ).all()


def build(rows, sc):
    now = datetime.now(timezone.utc).replace(tzinfo=None)
    out = []
    for r in rows:
        if classify(r.aggregated_status) in ("delivered", "returned", "cancelled"):
            continue
        sc_status = sc.get(r.order_number)
        vtext, vprio = verdict(sc_status)
        out.append(
            {
                "prio": vprio,
                "days": (now - r.frisbo_created_at).days,
                "row": [
                    r.order_number,
                    r.store,
                    r.aggregated_status,
                    classify(r.aggregated_status),
                    str(to_bucharest_date(r.frisbo_created_at) or ""),
                    (now - r.frisbo_created_at).days,
                    "Yes" if (r.tracking_number or "").strip() else "No",
                    round(float(r.total_price or 0), 2),
                    r.currency or "RON",
                    sc_status or "—",
                    vtext,
                    (r.note or "")[:200],
                ],
            }
        )
    # Actionable (STALE) first, then oldest first.
    out.sort(key=lambda x: (x["prio"], -x["days"]))
    return out


def _style_header(ws):
    fill = PatternFill("solid", fgColor="1F2937")
    font = Font(bold=True, color="FFFFFF")
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def _fill_sheet(ws, items):
    stale_fill = PatternFill("solid", fgColor="FECACA")  # red = stale (action)
    ws.append(HEADERS)
    _style_header(ws)
    for it in items:
        ws.append(it["row"])
        if it["prio"] == 0:  # FRISBO STALE
            for c in range(1, len(HEADERS) + 1):
                ws.cell(ws.max_row, c).fill = stale_fill
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"
    for i, w in enumerate([16, 18, 26, 11, 13, 9, 8, 10, 9, 22, 34, 36], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_xlsx(items, sc):
    wb = Workbook()
    _fill_sheet(wb.active, items)
    wb.active.title = "All Frozen"

    stale = [it for it in items if it["prio"] == 0]
    _fill_sheet(wb.create_sheet("Frisbo STALE (action)"), stale)

    s = wb.create_sheet("Summary")
    s.append(["FROZEN FRISBO ORDERS vs SCRIPTURI courier status"])
    s["A1"].font = Font(bold=True, size=13)
    s.append([f"Criterion: non-terminal Frisbo status AND older than {DAYS} days"])
    s.append([f"Scripturi source: {os.path.basename(SC_DB)} (snapshot)"])
    s.append(
        [f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"]
    )
    s.append(
        [
            f"Total frozen: {len(items)}  |  genuinely stale (courier resolved): {len(stale)}"
        ]
    )
    s.append([])
    for title, key in [
        ("By verdict", lambda it: it["row"][10]),
        ("By Scripturi status", lambda it: it["row"][9]),
        ("By Frisbo status", lambda it: it["row"][2]),
        ("By store", lambda it: it["row"][1]),
    ]:
        s.append([title, "Count"])
        s.cell(s.max_row, 1).font = Font(bold=True)
        s.cell(s.max_row, 2).font = Font(bold=True)
        for k, n in Counter(key(it) for it in items).most_common():
            s.append([k, n])
        s.append([])
    s.column_dimensions["A"].width = 38
    s.column_dimensions["B"].width = 12

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)


async def main():
    sc = load_scripturi()
    rows = await fetch_frozen()
    items = build(rows, sc)
    write_xlsx(items, sc)
    stale = sum(1 for it in items if it["prio"] == 0)
    print(
        f"Frozen orders: {len(items):,}  | genuinely stale (Scripturi shows resolved): {stale:,}"
    )
    print(f"WROTE {OUT}")


if __name__ == "__main__":
    asyncio.run(main())
