"""
FINAL list — ONLY the genuinely-stale Frisbo orders: those where Frisbo still shows
a NON-terminal aggregated_status but Scripturi's courier feed has already settled
them to a terminal state (Livrata / Anulata / Refuzata). Everything Scripturi also
shows as unresolved (Netrimisa / Lipsa awb / In curs) is excluded — those aren't
Frisbo errors.

"Stale" = Frisbo non-terminal AND older than `days` (default 14, so it's not just
normal courier lag) AND Scripturi status is terminal.

Run: cd backend && ./venv/Scripts/python.exe -u scratch/export_stale_orders_final.py [days] [out.xlsx] [scripturi.db]
Read-only on both sides.
"""

import os
import sys
import asyncio
import sqlite3
from collections import Counter, defaultdict
from datetime import datetime, timezone

sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, r"c:/Users/Admin/Desktop/AWB Print/awb-print-manager/backend")

from sqlalchemy import text  # noqa: E402
from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Font, PatternFill, Alignment  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

from app.core.database import AsyncSessionLocal  # noqa: E402
from app.core.status_classification import (  # noqa: E402
    classify,
    DELIVERED,
    RETURNED,
    REFUSED,
    CANCELLED,
)
from app.core.timezone import to_bucharest_date  # noqa: E402

DAYS = int(sys.argv[1]) if len(sys.argv) > 1 else 14
OUT = (
    sys.argv[2]
    if len(sys.argv) > 2
    else r"c:/Users/Admin/Desktop/stale_frisbo_orders_FINAL.xlsx"
)
SC_DB = (
    sys.argv[3]
    if len(sys.argv) > 3
    else r"c:/tmp/scr_new2/Scripturi/data/profitability.db"
)

TERMINAL = DELIVERED | RETURNED | REFUSED | CANCELLED
# Scripturi terminal status -> (courier outcome label, sort priority — Delivered first)
SC_TERMINAL = {
    "Livrata": ("Delivered", 0),
    "Refuzata": ("Returned/Refused", 1),
    "Anulata": ("Cancelled", 2),
}

HEADERS = [
    "Order #",
    "Store",
    "Frisbo Status (stale)",
    "Courier Outcome",
    "Scripturi Status",
    "Created (RO)",
    "Days Frozen",
    "Has AWB",
    "Total",
    "Currency",
    "Note",
]


def load_scripturi():
    con = sqlite3.connect(SC_DB)
    con.row_factory = sqlite3.Row
    sc = {
        r["order_name"]: r["status_category"]
        for r in con.execute("SELECT order_name, status_category FROM profit_orders")
    }
    con.close()
    return sc


async def fetch_frozen():
    sql = text(
        """
        SELECT o.order_number, COALESCE(s.name, o.store_uid) AS store,
               o.aggregated_status, o.tracking_number, o.frisbo_created_at,
               o.total_price, o.currency, o.note
        FROM orders o
        LEFT JOIN stores s ON s.uid = o.store_uid
        WHERE o.frisbo_created_at IS NOT NULL
          AND o.frisbo_created_at < (now() at time zone 'utc') - make_interval(days => :d)
          AND lower(coalesce(o.aggregated_status, '')) <> ALL(:term)
        ORDER BY o.frisbo_created_at ASC
        """
    )
    async with AsyncSessionLocal() as db:
        return (
            await db.execute(sql, {"d": DAYS, "term": [t.lower() for t in TERMINAL]})
        ).all()


def build(rows, sc):
    now = datetime.now(timezone.utc).replace(tzinfo=None)
    out = []
    for r in rows:
        if classify(r.aggregated_status) in ("delivered", "returned", "cancelled"):
            continue
        sc_status = sc.get(r.order_number)
        if sc_status not in SC_TERMINAL:
            continue  # keep ONLY the genuinely-stale (courier reached a terminal state)
        outcome, prio = SC_TERMINAL[sc_status]
        out.append(
            {
                "prio": prio,
                "days": (now - r.frisbo_created_at).days,
                "outcome": outcome,
                "currency": r.currency or "RON",
                "total": round(float(r.total_price or 0), 2),
                "row": [
                    r.order_number,
                    r.store,
                    r.aggregated_status,
                    outcome,
                    sc_status,
                    str(to_bucharest_date(r.frisbo_created_at) or ""),
                    (now - r.frisbo_created_at).days,
                    "Yes" if (r.tracking_number or "").strip() else "No",
                    round(float(r.total_price or 0), 2),
                    r.currency or "RON",
                    (r.note or "")[:200],
                ],
            }
        )
    out.sort(key=lambda x: (x["prio"], -x["days"]))  # Delivered first, oldest first
    return out


def write_xlsx(items):
    wb = Workbook()
    ws = wb.active
    ws.title = "Stale Orders"
    ws.append(HEADERS)
    hf, hfont = PatternFill("solid", fgColor="1F2937"), Font(bold=True, color="FFFFFF")
    for c in range(1, len(HEADERS) + 1):
        ws.cell(1, c).fill = hf
        ws.cell(1, c).font = hfont
        ws.cell(1, c).alignment = Alignment(horizontal="center")
    out_fill = {
        "Delivered": PatternFill("solid", fgColor="DCFCE7"),
        "Returned/Refused": PatternFill("solid", fgColor="FEE2E2"),
        "Cancelled": PatternFill("solid", fgColor="FEF3C7"),
    }
    for it in items:
        ws.append(it["row"])
        ws.cell(ws.max_row, 4).fill = out_fill[it["outcome"]]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"
    for i, w in enumerate([16, 18, 24, 17, 20, 13, 12, 8, 10, 9, 36], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Summary
    s = wb.create_sheet("Summary")
    s.append(["GENUINELY-STALE FRISBO ORDERS (courier already settled them)"])
    s["A1"].font = Font(bold=True, size=13)
    s.append(
        [
            f"Criterion: Frisbo non-terminal + older than {DAYS}d + Scripturi status terminal"
        ]
    )
    s.append([f"Scripturi source: {os.path.basename(SC_DB)} (snapshot)"])
    s.append(
        [f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"]
    )
    s.append([f"TOTAL stale orders: {len(items)}"])
    s.append([])
    s.append(["By courier outcome", "Count"])
    s.cell(s.max_row, 1).font = s.cell(s.max_row, 2).font = Font(bold=True)
    for k, n in Counter(it["outcome"] for it in items).most_common():
        s.append([k, n])
    s.append([])
    # Under-counted DELIVERED value, per currency (AWB never booked these as delivered)
    s.append(["Delivered-but-frozen value (per currency)", "Sum"])
    s.cell(s.max_row, 1).font = s.cell(s.max_row, 2).font = Font(bold=True)
    by_ccy = defaultdict(float)
    for it in items:
        if it["outcome"] == "Delivered":
            by_ccy[it["currency"]] += it["total"]
    for ccy, v in sorted(by_ccy.items(), key=lambda x: -x[1]):
        s.append([ccy, round(v, 2)])
    s.append([])
    s.append(["By store", "Count"])
    s.cell(s.max_row, 1).font = s.cell(s.max_row, 2).font = Font(bold=True)
    for k, n in Counter(it["row"][1] for it in items).most_common():
        s.append([k, n])
    s.append([])
    s.append(["By Frisbo (stale) status", "Count"])
    s.cell(s.max_row, 1).font = s.cell(s.max_row, 2).font = Font(bold=True)
    for k, n in Counter(it["row"][2] for it in items).most_common():
        s.append([k, n])
    s.column_dimensions["A"].width = 42
    s.column_dimensions["B"].width = 14

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)


async def main():
    sc = load_scripturi()
    rows = await fetch_frozen()
    items = build(rows, sc)
    write_xlsx(items)
    bo = Counter(it["outcome"] for it in items)
    print(f"Stale orders: {len(items):,}  ({dict(bo)})")
    print(f"WROTE {OUT}")


if __name__ == "__main__":
    asyncio.run(main())
