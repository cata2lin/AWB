"""
DPD Weight-Audit — recoverable-money dashboard.

Ported from Scripturi's ``dpd_audit.py`` two-pass weight-learning engine, but
adapted to AWB's own data instead of Scripturi's JSON nomenclator + Shopify:

  • Real billed weight comes from ``OrderAwb.package_weight`` (parsed from courier
    CSV imports, see ``app/api/courier_csv/parsers.py`` — DPD ``Greutate`` column).
  • Per-AWB cost comes from ``OrderAwb.transport_cost`` (cu TVA), already RON.
  • The AWB → order → SKU mapping is AWB's own: each ``OrderAwb`` joins to its
    ``Order`` (by ``order_id``), whose ``line_items`` carry SKU + quantity.

Two endpoints:

  POST /api/analytics/courier-audit         — analyze an UPLOADED DPD invoice file
       (xls/xlsx/csv). Reuses the proven Scripturi description-parser + dim parser.
       This is the "fresh invoice, find overcharges before you pay" flow and is the
       only path that can compute true VOLUMETRIC status (the invoice carries dims).

  GET  /api/analytics/courier-audit/imported — analyze ALREADY-imported courier_csv
       data for a date range (no upload). Uses the AWB→order→SKU mapping to learn
       per-SKU median billed weight and flag AWBs whose billed weight is well above
       what their SKUs historically weigh. AWB has no product dimensions, so this
       path does the weight-vs-learned-median analysis and reports that volumetric
       needs product dims.

Two-pass logic (both paths):
  PASS 1 — learn a robust per-SKU "real" weight: collect per-unit billed weights
           from single-SKU shipments, reject outliers (>2× median), take the median.
  PASS 2 — classify each AWB:
           OK         billed ≈ expected (within threshold)
           PROBLEM    billed well above expected, NOT explained by volumetric
           VOLUMETRIC billed above physical weight but explained by L*W*H/5000
           PARTIAL    at least one SKU unknown (no learned/known weight)
           UNDER      billed below expected (informational)
           NO_SKU     no SKUs could be resolved for the AWB

Read-only. No writes to any table.
"""

from __future__ import annotations

import csv
import io
import re
import statistics
from typing import Optional, Any

from fastapi import APIRouter, Depends, UploadFile, File, Form, HTTPException
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.core.analytics_cache import cached_analytics
from app.core.timezone import date_str_to_utc_start, date_str_to_utc_end
from app.models import Order, OrderAwb
from app.models.order_awb import is_billable_status

router = APIRouter()

VOLUMETRIC_DIVISOR = 5000.0


# ────────────────────────────────────────────────────────────────
# Generic value parsing (mirrors Scripturi dpd_audit.py helpers)
# ────────────────────────────────────────────────────────────────

DIACRITICS = str.maketrans(
    {
        "ă": "a",
        "â": "a",
        "î": "i",
        "ș": "s",
        "ş": "s",
        "ț": "t",
        "ţ": "t",
        "Ă": "a",
        "Â": "a",
        "Î": "i",
        "Ș": "s",
        "Ş": "s",
        "Ț": "t",
        "Ţ": "t",
    }
)


def _norm_text(value: Any) -> str:
    s = str(value or "").strip().translate(DIACRITICS).lower()
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def _clean_sku(value: Any) -> str:
    s = str(value or "").strip().upper().replace(" ", " ")
    s = re.sub(r"\s+", " ", s)
    return s.strip(" \t\r\n,;")


def _parse_float(value: Any, default: float = 0.0) -> float:
    if value is None:
        return default
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value != value:  # NaN
            return default
        return float(value)
    s = str(value).strip()
    if not s:
        return default
    s = s.replace(" ", " ").replace('"', "")
    for token in ("RON", "ron", "LEI", "lei", "KG", "kg", "CM", "cm"):
        s = s.replace(token, "")
    m = re.search(r"[-+]?\d+(?:[.,]\d+)?", s.strip())
    if not m:
        return default
    try:
        return float(m.group(0).replace(",", "."))
    except ValueError:
        return default


# ────────────────────────────────────────────────────────────────
# Description / SKU parser (ported from Scripturi parse_description)
# ────────────────────────────────────────────────────────────────

ITEM_PATTERN = re.compile(
    r"(\d+)\s*[Xx]\s*([A-Za-z0-9][A-Za-z0-9_\-\. ]*?)(?=\s*(?:\||/|,|\+|$))",
    re.IGNORECASE,
)


def _parse_description(desc: str) -> list[dict[str, Any]]:
    """Parse a DPD parcel content string into [{sku, qty}]."""
    if not desc:
        return []
    text = re.sub(r"\s+", " ", str(desc).replace("\n", " ").replace("\r", " ").strip())

    if "/" in text:
        before, after = text.split("/", 1)
        scan_text = (before + " | " + after) if ITEM_PATTERN.search(before) else after
    else:
        scan_text = text
    scan_text = scan_text.strip()

    if "|" in scan_text or " + " in scan_text:
        sku = _clean_sku(scan_text)
        return [{"sku": sku, "qty": 1}] if sku else []

    items: list[dict[str, Any]] = []
    for qty_str, raw_sku in ITEM_PATTERN.findall(scan_text):
        sku = _clean_sku(raw_sku)
        if not sku or sku in {"X", "KG", "CM"}:
            continue
        items.append({"sku": sku, "qty": int(qty_str)})

    if not items and scan_text:
        sku = _clean_sku(scan_text)
        if sku and not sku.isnumeric():
            items.append({"sku": sku, "qty": 1})
    return items


# ────────────────────────────────────────────────────────────────
# Dimensions / volumetric (ported)
# ────────────────────────────────────────────────────────────────

DIM_PATTERN = re.compile(
    r"(?:(\d+(?:[.,]\d+)?)\s*[xX]\s*)?"
    r"(\d+(?:[.,]\d+)?)\s*[xX]\s*"
    r"(\d+(?:[.,]\d+)?)\s*[xX]\s*"
    r"(\d+(?:[.,]\d+)?)",
    re.IGNORECASE,
)


def _parse_dimensions(text: Any) -> list[dict[str, float]]:
    if not text:
        return []
    s = re.sub(
        r"\s+", " ", str(text).replace("×", "x").replace("*", "x").replace(" ", " ")
    )
    groups: list[dict[str, float]] = []
    for match in DIM_PATTERN.finditer(s):
        qty_raw, l_raw, w_raw, h_raw = match.groups()
        qty = _parse_float(qty_raw, 1.0) if qty_raw else 1.0
        l, w, h = _parse_float(l_raw), _parse_float(w_raw), _parse_float(h_raw)
        if l <= 0 or w <= 0 or h <= 0:
            continue
        groups.append({"qty": qty, "lungime": l, "latime": w, "inaltime": h})
    return groups


def _volumetric_kg(
    l: float, w: float, h: float, divisor: float = VOLUMETRIC_DIVISOR
) -> float:
    if not l or not w or not h:
        return 0.0
    return round((float(l) * float(w) * float(h)) / float(divisor), 3)


# ────────────────────────────────────────────────────────────────
# Uploaded-file readers (csv / xlsx / xls)
# ────────────────────────────────────────────────────────────────


def _cell_to_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _read_csv_bytes(raw: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    text = raw.decode("utf-8-sig", errors="replace")
    first_line = text.splitlines()[0] if text.splitlines() else ""
    if "\t" in first_line:
        delimiter = "\t"
    elif ";" in first_line and ";" in first_line:
        delimiter = ";"
    else:
        try:
            delimiter = csv.Sniffer().sniff(text[:4096], delimiters=",;\t").delimiter
        except Exception:
            delimiter = ","
    rows_raw = [
        r
        for r in csv.reader(io.StringIO(text), delimiter=delimiter)
        if any(str(c).strip() for c in r)
    ]
    if not rows_raw:
        return [], []
    headers = [_cell_to_str(h) or f"col{i + 1}" for i, h in enumerate(rows_raw[0])]
    out_rows = [
        {h: (raw_row[i] if i < len(raw_row) else "") for i, h in enumerate(headers)}
        for raw_row in rows_raw[1:]
    ]
    return headers, out_rows


def _read_xlsx_bytes(raw: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise HTTPException(
            500, "openpyxl indisponibil pe server pentru .xlsx"
        ) from exc
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)
    first = next(rows_iter, None)
    if not first:
        return [], []
    headers = [_cell_to_str(h) or f"col{i + 1}" for i, h in enumerate(first)]
    rows = []
    for values in rows_iter:
        row = {
            h: _cell_to_str(values[i] if i < len(values) else "")
            for i, h in enumerate(headers)
        }
        if any(str(v).strip() for v in row.values()):
            rows.append(row)
    return headers, rows


def _read_xls_bytes(raw: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover
        raise HTTPException(500, "xlrd indisponibil pe server pentru .xls") from exc
    wb = xlrd.open_workbook(file_contents=raw)
    ws = wb.sheet_by_index(0)
    if ws.nrows == 0:
        return [], []
    headers = [
        _cell_to_str(ws.cell_value(0, c)) or f"col{c + 1}" for c in range(ws.ncols)
    ]
    rows = []
    for r in range(1, ws.nrows):
        rows.append(
            {h: _cell_to_str(ws.cell_value(r, c)) for c, h in enumerate(headers)}
        )
    return headers, rows


def _read_upload(filename: str, raw: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    name = (filename or "").lower()
    if name.endswith(".csv"):
        return _read_csv_bytes(raw)
    if name.endswith(".xlsx"):
        return _read_xlsx_bytes(raw)
    if name.endswith(".xls"):
        return _read_xls_bytes(raw)
    raise HTTPException(
        400, f"Format nesuportat. Folosește .csv, .xls sau .xlsx (primit: {filename})"
    )


def _detect_columns(headers: list[str]) -> dict[str, str]:
    col: dict[str, str] = {}
    for h in headers:
        n = _norm_text(h)
        if not n:
            continue
        if ("awb" in n or "waybill" in n or n == "expediere") and "awb" not in col:
            col["awb"] = h
            continue
        if n in {"data", "date"} and "date" not in col:
            col["date"] = h
            continue
        if (
            "greutate corecta" in n or "correct weight" in n
        ) and "correct_weight" not in col:
            col["correct_weight"] = h
            continue
        if (
            n in {"weight", "greutate"}
            or "greutate colet" in n
            or "billed weight" in n
            or "charged weight" in n
        ) and "weight" not in col:
            col["weight"] = h
            continue
        if (
            "descriere" in n or "description" in n or "continut" in n
        ) and "description" not in col:
            col["description"] = h
            continue
        if ("dimensiune" in n or "dimension" in n) and "dimensions" not in col:
            col["dimensions"] = h
            continue
        if n.endswith(" id") or n.split()[-1:] == ["id"]:
            continue
        if "owed amount" in n and "amount" not in col:
            col["amount"] = h
            continue
        cost_terms = (
            "total transport",
            "cost transport",
            "pret transport",
            "valoare transport",
            "total fara tva",
        )
        if (
            any(t in n for t in cost_terms) or n in {"amount", "cost", "tarif", "total"}
        ) and "amount" not in col:
            col["amount"] = h
            continue
    return col


def _parse_uploaded_invoice(filename: str, raw: bytes) -> list[dict[str, Any]]:
    headers, rows = _read_upload(filename, raw)
    col = _detect_columns(headers)
    if "awb" not in col or "weight" not in col:
        raise HTTPException(
            400,
            "Nu am găsit coloanele de bază (AWB + Greutate) în fișier. "
            "Verifică să fie un export/factură DPD cu antet.",
        )
    parcels: list[dict[str, Any]] = []
    for row in rows:
        awb = re.sub(r"\.0$", "", _cell_to_str(row.get(col["awb"], "")).strip())
        if not awb:
            continue
        weight = _parse_float(row.get(col["weight"]))
        if weight <= 0:
            continue
        dims_text = _cell_to_str(row.get(col.get("dimensions", ""), ""))
        parcels.append(
            {
                "awb": awb,
                "date": _cell_to_str(row.get(col.get("date", ""), "")),
                "weight_dpd": round(weight, 3),
                "description": _cell_to_str(row.get(col.get("description", ""), "")),
                "amount": round(_parse_float(row.get(col.get("amount", ""))), 2),
                "dimensions_input": dims_text,
                "dimensions_groups_input": _parse_dimensions(dims_text),
            }
        )
    return parcels


# ────────────────────────────────────────────────────────────────
# Two-pass analysis engine (shared between both flows)
# ────────────────────────────────────────────────────────────────


def _learn_reference_weights(parcels: list[dict[str, Any]]) -> dict[str, float]:
    """PASS 1: per-SKU robust per-unit weight from single-SKU parcels."""
    samples: dict[str, list[float]] = {}
    for parcel in parcels:
        items = parcel.get("items") or _parse_description(parcel.get("description", ""))
        if len(items) != 1:
            continue
        sku = items[0]["sku"]
        qty = max(int(items[0].get("qty") or 1), 1)
        w = _parse_float(parcel.get("weight_dpd"))
        if w <= 0:
            continue
        per_unit = round(w / qty, 3)
        existing = samples.get(sku, [])
        if len(existing) >= 3:
            med = statistics.median(existing)
            if med > 0 and per_unit > 3 * med:  # gross outlier — skip
                continue
        samples.setdefault(sku, []).append(per_unit)

    ref: dict[str, float] = {}
    for sku, vals in samples.items():
        if len(vals) >= 3:
            med = statistics.median(vals)
            filtered = [v for v in vals if v <= 2 * med] or [med]
            ref[sku] = round(statistics.median(filtered), 3)
        elif vals:
            ref[sku] = round(statistics.median(vals), 3)
    return ref


def _analyze(
    parcels: list[dict[str, Any]],
    threshold_kg: float,
    min_weight_kg: float,
    has_dims: bool,
) -> dict[str, Any]:
    learned_ref = _learn_reference_weights(parcels)

    results: list[dict[str, Any]] = []
    unknown_skus: set[str] = set()
    problems = volumetric_explained = 0
    total_excess_kg = total_estimated_excess_cost = 0.0
    total_amount = total_billed_weight = total_expected = 0.0

    for parcel in parcels:
        weight_dpd = _parse_float(parcel.get("weight_dpd"))
        total_amount += _parse_float(parcel.get("amount"))
        total_billed_weight += weight_dpd
        if weight_dpd <= min_weight_kg:
            continue

        items = parcel.get("items") or _parse_description(parcel.get("description", ""))
        dims_groups = parcel.get("dimensions_groups_input") or _parse_dimensions(
            parcel.get("dimensions_input", "")
        )

        base = {
            "awb": parcel.get("awb", ""),
            "date": parcel.get("date", ""),
            "amount": round(_parse_float(parcel.get("amount")), 2),
            "weight_dpd": round(weight_dpd, 3),
            "description": parcel.get("description", ""),
            "items": items,
            "dimensions": parcel.get("dimensions_input", ""),
            "order_ref": parcel.get("order_ref", ""),
            "store_name": parcel.get("store_name", ""),
        }

        if not items:
            results.append(
                {
                    **base,
                    "weight_calculated": None,
                    "volumetric_weight": None,
                    "expected_billable_weight": None,
                    "difference": None,
                    "billable_difference": None,
                    "status": "NO_SKU",
                    "status_label": "Fără SKU",
                    "missing_skus": [],
                }
            )
            continue

        calculated = 0.0
        all_found = True
        missing: list[str] = []
        for item in items:
            sku, qty = item["sku"], int(item.get("qty") or 1)
            unit = learned_ref.get(sku, 0.0)
            if unit > 0:
                calculated += qty * unit
            else:
                all_found = False
                missing.append(sku)
                unknown_skus.add(sku)
        calculated = round(calculated, 3)

        volumetric = 0.0
        for dg in dims_groups:
            volumetric += _parse_float(dg.get("qty"), 1.0) * _volumetric_kg(
                dg.get("lungime", 0), dg.get("latime", 0), dg.get("inaltime", 0)
            )
        volumetric = round(volumetric, 3)

        expected = calculated if all_found else None
        total_expected += expected or 0
        billable_diff = (
            round(weight_dpd - expected, 3) if expected is not None else None
        )

        if not all_found:
            status, label = "PARTIAL", "SKU lipsă"
        elif billable_diff is not None and billable_diff > threshold_kg:
            if volumetric > calculated and weight_dpd <= volumetric + threshold_kg:
                status, label = "VOLUMETRIC", "Volumetric"
                volumetric_explained += 1
            else:
                status, label = "PROBLEM", "Problemă"
            problems += 1
            total_excess_kg += billable_diff
            amount = _parse_float(parcel.get("amount"))
            if amount > 0 and weight_dpd > 0:
                total_estimated_excess_cost += amount * (billable_diff / weight_dpd)
        elif billable_diff is not None and billable_diff < -threshold_kg:
            status, label = "UNDER", "Sub greutate"
        else:
            status, label = "OK", "OK"

        results.append(
            {
                **base,
                "weight_calculated": calculated if all_found else None,
                "volumetric_weight": volumetric if volumetric > 0 else None,
                "expected_billable_weight": expected,
                "difference": billable_diff,
                "billable_difference": billable_diff,
                "status": status,
                "status_label": label,
                "missing_skus": missing,
            }
        )

    order = {
        "PROBLEM": 0,
        "PARTIAL": 1,
        "VOLUMETRIC": 2,
        "UNDER": 3,
        "OK": 4,
        "NO_SKU": 5,
    }
    results.sort(
        key=lambda r: (
            order.get(r["status"], 9),
            -(r.get("billable_difference") or 0),
        )
    )

    summary = {
        "total_parcels": len(parcels),
        "parcels_above_min_weight": len(results),
        "min_weight_kg": min_weight_kg,
        "threshold_kg": threshold_kg,
        "problems": problems,
        "volumetric_explained": volumetric_explained,
        "partial": sum(1 for r in results if r["status"] == "PARTIAL"),
        "ok": sum(1 for r in results if r["status"] == "OK"),
        "total_excess_kg": round(total_excess_kg, 2),
        "estimated_excess_cost": round(total_estimated_excess_cost, 2),
        "total_amount": round(total_amount, 2),
        "total_billed_weight": round(total_billed_weight, 2),
        "total_expected_billable_weight": round(total_expected, 2),
        "unknown_skus": sorted(unknown_skus),
        "unknown_skus_count": len(unknown_skus),
        "learned_skus_count": len(learned_ref),
        "has_dimensions": has_dims,
    }
    if not has_dims:
        summary["volumetric_note"] = (
            "AWB nu stochează dimensiunile produselor, deci nu se poate calcula "
            "greutatea volumetrică. Analiza compară greutatea facturată cu mediana "
            "învățată per SKU. Pentru detecția volumetrică, încarcă o factură DPD cu "
            "coloana Dimensiuni sau adaugă dimensiuni la produse."
        )

    # Dispute-ready subset: only billable overcharges (PROBLEM), with the fields a
    # DPD complaint needs.
    dispute_rows = [
        {
            "awb": r["awb"],
            "date": r["date"],
            "order_ref": r.get("order_ref", ""),
            "billed_weight": r["weight_dpd"],
            "expected_weight": r["expected_billable_weight"],
            "excess_kg": r["billable_difference"],
            "amount": r["amount"],
            "estimated_excess_cost": (
                round(r["amount"] * (r["billable_difference"] / r["weight_dpd"]), 2)
                if r["amount"] and r["weight_dpd"]
                else 0.0
            ),
            "description": r["description"],
        }
        for r in results
        if r["status"] == "PROBLEM"
    ]

    return {"results": results, "summary": summary, "dispute_rows": dispute_rows}


# ────────────────────────────────────────────────────────────────
# Endpoint 1 — analyze an UPLOADED DPD invoice file
# ────────────────────────────────────────────────────────────────


@router.post("/analytics/courier-audit")
async def courier_audit_upload(
    file: UploadFile = File(...),
    threshold: float = Form(0.5),
    min_weight: float = Form(3.0),
    db: AsyncSession = Depends(get_db),
):
    """Analyze an uploaded DPD invoice/export (.csv/.xls/.xlsx) for weight overcharges.

    This path can compute VOLUMETRIC status when the file carries a Dimensiuni column.
    """
    raw = await file.read()
    if not raw:
        raise HTTPException(400, "Fișier gol.")
    if len(raw) > 25 * 1024 * 1024:
        raise HTTPException(413, "Fișier prea mare (max 25 MB).")

    parcels = _parse_uploaded_invoice(file.filename or "upload", raw)
    if not parcels:
        raise HTTPException(400, "Niciun colet valid găsit în fișier.")

    has_dims = any(p.get("dimensions_groups_input") for p in parcels)
    out = _analyze(parcels, max(0.0, threshold), max(0.0, min_weight), has_dims)
    out["source"] = "upload"
    out["filename"] = file.filename
    return out


# ────────────────────────────────────────────────────────────────
# Endpoint 2 — analyze ALREADY-imported courier_csv data for a range
# ────────────────────────────────────────────────────────────────


def _order_items(order: Optional[Order]) -> list[dict[str, Any]]:
    """Extract [{sku, qty}] from an Order's line_items (AWB→order→SKU mapping)."""
    if not order or not isinstance(order.line_items, list):
        return []
    items: list[dict[str, Any]] = []
    for li in order.line_items:
        if not isinstance(li, dict):
            continue
        inv = li.get("inventory_item") or {}
        sku = _clean_sku(inv.get("sku") if isinstance(inv, dict) else None)
        if not sku:
            continue
        qty = li.get("quantity", 1) or 1
        try:
            qty = int(qty)
        except (ValueError, TypeError):
            qty = 1
        items.append({"sku": sku, "qty": max(qty, 1)})
    return items


@router.get("/analytics/courier-audit/imported")
@cached_analytics("courier-audit-imported")
async def courier_audit_imported(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    store_uids: Optional[str] = None,
    courier: Optional[str] = "DPD",
    threshold: float = 0.5,
    min_weight: float = 3.0,
    db: AsyncSession = Depends(get_db),
):
    """Analyze already-imported courier AWB data (OrderAwb.package_weight / cost) for a range.

    Uses AWB's own AWB→order→SKU mapping; learns per-SKU median billed weight and
    flags AWBs billed well above their SKUs' learned weight. No dims in AWB → no
    volumetric detection (reported in summary.volumetric_note).
    """
    # Date range bounds the scan. Without it, this would load ALL-TIME courier AWBs
    # (order_awbs has 600k+ rows) — default to the last 30 days. The filter is pushed
    # to SQL (NOT applied in Python after loading everything, which is what made the
    # original load the whole table and then crash asyncpg on a huge IN-clause).
    if date_from and date_to:
        start = date_str_to_utc_start(date_from)
        end = date_str_to_utc_end(date_to)
    else:
        from datetime import timedelta
        from app.core.timezone import romania_now, UTC_TZ

        _now = romania_now()
        start = (_now - timedelta(days=30)).astimezone(UTC_TZ).replace(tzinfo=None)
        end = _now.astimezone(UTC_TZ).replace(tzinfo=None)

    conditions = [
        OrderAwb.package_weight.isnot(None),
        OrderAwb.awb_type == "outbound",
        Order.frisbo_created_at >= start,
        Order.frisbo_created_at <= end,
    ]
    if courier:
        conditions.append(OrderAwb.courier_name.ilike(f"%{courier.strip()}%"))
    if store_uids:
        store_filter = {s.strip() for s in store_uids.split(",") if s.strip()}
        if store_filter:
            conditions.append(Order.store_uid.in_(store_filter))

    # Single bounded JOIN (OrderAwb -> Order) — replaces the unbounded all-time scan
    # + the Order.id.in_(tens-of-thousands) that segfaulted asyncpg. Each row is an
    # (OrderAwb, Order) tuple.
    rows = (
        await db.execute(
            select(OrderAwb, Order)
            .join(Order, OrderAwb.order_id == Order.id)
            .where(*conditions)
        )
    ).all()

    parcels: list[dict[str, Any]] = []
    skipped_non_billable = 0
    for a, order in rows:
        if not is_billable_status(a.csv_status):
            skipped_non_billable += 1
            continue
        ref_date = (
            (order.frisbo_created_at if order else None)
            or a.shipment_created_at
            or a.created_at
        )

        weight = _parse_float(a.package_weight)
        if weight <= 0:
            continue
        parcels.append(
            {
                "awb": a.tracking_number or "",
                "date": ref_date.strftime("%Y-%m-%d") if ref_date else "",
                "weight_dpd": round(weight, 3),
                "amount": round(_parse_float(a.transport_cost), 2),
                "items": _order_items(order),
                "order_ref": a.order_ref or (order.order_number if order else ""),
                "store_name": (order.store_uid if order else ""),
                "description": "",
                "dimensions_input": "",
                "dimensions_groups_input": [],
            }
        )

    out = _analyze(parcels, max(0.0, threshold), max(0.0, min_weight), has_dims=False)
    out["source"] = "imported"
    out["courier"] = courier
    out["summary"]["awbs_scanned"] = len(rows)
    out["summary"]["skipped_non_billable"] = skipped_non_billable
    return out
