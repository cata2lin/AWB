"""
Products API endpoints — grouped inventory view, exclusion, and Excel export.

Grouping logic: products are grouped by BARCODE first, then by SKU for any
remaining products. This means a product with SKU "10" and barcode "786..."
will be grouped with another product that has SKU "10" but no barcode.
"""
from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, or_, cast, String, update
from typing import Optional, List
import io

from app.core.database import get_db
from app.models.product import Product
from app.models.store import Store

router = APIRouter()


# ── Helpers ──

def _product_to_listing(p, store_names):
    """Convert a Product ORM object to a listing dict."""
    return {
        "uid": p.uid,
        "title_1": p.title_1,
        "title_2": p.title_2,
        "sku": p.sku,
        "barcode": p.barcode,
        "state": p.state,
        "images": p.images or [],
        "store_uids": p.store_uids or [],
        "stores": [
            {"uid": uid, "name": store_names.get(uid, uid)}
            for uid in (p.store_uids or [])
        ],
        "stock_available": p.stock_available,
        "stock_committed": p.stock_committed,
        "stock_incoming": p.stock_incoming,
        "synced_at": p.synced_at,
        "frisbo_updated_at": p.frisbo_updated_at,
        "organization_uid": p.organization_uid,
        "missing_barcode": not (p.barcode or "").strip(),
    }


def _build_groups(all_products):
    """
    Group products by barcode OR SKU. Priority:
    1. If a product has a barcode, group by barcode
    2. If no barcode but has SKU, try to join an existing barcode group with same SKU
    3. If no barcode and SKU doesn't match any barcode group, group by SKU alone
    4. Products with neither barcode nor SKU remain ungrouped
    """
    # Phase 1: group products with barcodes
    barcode_groups = {}  # barcode -> [products]
    sku_to_barcode = {}  # sku -> barcode (maps SKUs to their barcode group)
    remaining = []       # products without barcode

    for p in all_products:
        bc = (p.barcode or "").strip()
        sku = (p.sku or "").strip()

        if bc:
            barcode_groups.setdefault(bc, []).append(p)
            # Track all SKUs that belong to this barcode group
            if sku:
                sku_to_barcode[sku] = bc
        else:
            remaining.append(p)

    # Phase 2: try to merge remaining products into barcode groups by SKU
    sku_only_groups = {}  # sku -> [products] (for those that don't match any barcode)
    ungrouped = []

    for p in remaining:
        sku = (p.sku or "").strip()
        if sku and sku in sku_to_barcode:
            # This SKU matches a barcode group — merge in
            bc = sku_to_barcode[sku]
            barcode_groups[bc].append(p)
        elif sku:
            # Group by SKU alone
            sku_only_groups.setdefault(sku, []).append(p)
        else:
            ungrouped.append(p)

    return barcode_groups, sku_only_groups, ungrouped


def _merge_group(group, group_key, store_names, sku_cost_map):
    """Merge a list of product listings into one grouped result dict."""
    # Sort: most recently synced first
    group.sort(
        key=lambda p: p.synced_at or p.frisbo_updated_at or p.frisbo_created_at or p.synced_at,
        reverse=True,
    )

    # Check if a DB-stored primary preference exists
    primary = group[0]
    stored_primary_uid = None
    for p in group:
        if p.primary_listing_uid:
            stored_primary_uid = p.primary_listing_uid
            break

    if stored_primary_uid:
        for p in group:
            if p.uid == stored_primary_uid:
                primary = p
                break

    # Merge stores from all listings
    all_store_uids = []
    seen = set()
    for p in group:
        for uid in (p.store_uids or []):
            if uid not in seen:
                all_store_uids.append(uid)
                seen.add(uid)

    # Best image (first listing that has one)
    images = []
    for p in group:
        if p.images:
            images = p.images
            break

    # Best title
    title_1, title_2 = primary.title_1, primary.title_2
    for p in group:
        if p.title_1:
            title_1, title_2 = p.title_1, p.title_2
            break

    # Cost from SKU costs table
    cost = None
    for p in group:
        if p.sku and p.sku in sku_cost_map:
            cost = sku_cost_map[p.sku]
            break

    # Check if any listing is missing barcode
    has_missing_barcode = any(not (p.barcode or "").strip() for p in group)

    return {
        "uid": primary.uid,
        "id": primary.id,
        "barcode": primary.barcode or group_key,
        "title_1": title_1,
        "title_2": title_2,
        "sku": primary.sku,
        "state": primary.state,
        "images": images,
        "store_uids": all_store_uids,
        "stores": [{"uid": uid, "name": store_names.get(uid, uid)} for uid in all_store_uids],
        "stock_available": primary.stock_available,
        "stock_committed": primary.stock_committed,
        "stock_incoming": primary.stock_incoming,
        "exclude_from_stock": primary.exclude_from_stock,
        "cost": cost,
        "grouped_count": len(group),
        "grouped_uids": [p.uid for p in group],
        "primary_uid": primary.uid,
        "has_missing_barcode": has_missing_barcode,
        "listings": [_product_to_listing(p, store_names) for p in group],
        "synced_at": primary.synced_at,
        "frisbo_updated_at": primary.frisbo_updated_at,
    }


async def _fetch_and_group(db, search, store_uid, state, has_stock, has_cost,
                            exclude_filter, missing_barcode_filter, sort_field,
                            sort_direction, skip=None, limit=None):
    """Shared logic for grouped list and Excel export."""
    query = select(Product)

    if search:
        sp = f"%{search}%"
        query = query.where(or_(
            Product.title_1.ilike(sp),
            Product.title_2.ilike(sp),
            Product.sku.ilike(sp),
            Product.barcode.ilike(sp),
        ))
    if state:
        query = query.where(Product.state == state)
    if has_stock is True:
        query = query.where(Product.stock_available > 0)
    elif has_stock is False:
        query = query.where(Product.stock_available == 0)
    if store_uid:
        query = query.where(cast(Product.store_uids, String).contains(store_uid))
    if exclude_filter == 'excluded':
        query = query.where(Product.exclude_from_stock == True)
    elif exclude_filter == 'active':
        query = query.where(Product.exclude_from_stock == False)

    result = await db.execute(query.order_by(Product.synced_at.desc()))
    all_products = result.scalars().all()

    # Load SKU costs
    from app.models.sku_cost import SkuCost
    costs_result = await db.execute(select(SkuCost.sku, SkuCost.cost))
    sku_cost_map = {row[0]: float(row[1]) for row in costs_result.all() if row[1] is not None}

    # Resolve store names
    all_store_uid_set = set()
    for p in all_products:
        for uid in (p.store_uids or []):
            all_store_uid_set.add(uid)

    store_names = {}
    if all_store_uid_set:
        stores_result = await db.execute(
            select(Store.uid, Store.name).where(Store.uid.in_(list(all_store_uid_set)))
        )
        store_names = {row[0]: row[1] for row in stores_result.all()}

    # Group
    barcode_groups, sku_only_groups, ungrouped = _build_groups(all_products)

    merged = []

    for bc, group in barcode_groups.items():
        merged.append(_merge_group(group, bc, store_names, sku_cost_map))

    for sku, group in sku_only_groups.items():
        merged.append(_merge_group(group, sku, store_names, sku_cost_map))

    for p in ungrouped:
        cost = sku_cost_map.get(p.sku) if p.sku else None
        merged.append({
            "uid": p.uid,
            "id": p.id,
            "barcode": p.barcode,
            "title_1": p.title_1,
            "title_2": p.title_2,
            "sku": p.sku,
            "state": p.state,
            "images": p.images,
            "store_uids": p.store_uids or [],
            "stores": [{"uid": uid, "name": store_names.get(uid, uid)} for uid in (p.store_uids or [])],
            "stock_available": p.stock_available,
            "stock_committed": p.stock_committed,
            "stock_incoming": p.stock_incoming,
            "exclude_from_stock": p.exclude_from_stock,
            "cost": cost,
            "grouped_count": 1,
            "grouped_uids": [p.uid],
            "primary_uid": p.uid,
            "has_missing_barcode": not (p.barcode or "").strip(),
            "listings": [_product_to_listing(p, store_names)],
            "synced_at": p.synced_at,
            "frisbo_updated_at": p.frisbo_updated_at,
        })

    # Post-grouping filters
    if has_cost == 'yes':
        merged = [m for m in merged if m.get("cost") is not None]
    elif has_cost == 'no':
        merged = [m for m in merged if m.get("cost") is None]

    if missing_barcode_filter == 'yes':
        merged = [m for m in merged if m.get("has_missing_barcode")]
    elif missing_barcode_filter == 'no':
        merged = [m for m in merged if not m.get("has_missing_barcode")]

    # Sort
    sort_key_map = {
        "title_1": lambda x: (x.get("title_1") or "").lower(),
        "sku": lambda x: (x.get("sku") or "").lower(),
        "stock_available": lambda x: x.get("stock_available", 0),
        "stock_committed": lambda x: x.get("stock_committed", 0),
        "barcode": lambda x: (x.get("barcode") or "").lower(),
        "cost": lambda x: x.get("cost") or 0,
        "synced_at": lambda x: str(x.get("synced_at") or ""),
        "grouped_count": lambda x: x.get("grouped_count", 1),
    }
    sort_fn = sort_key_map.get(sort_field, sort_key_map["title_1"])
    merged.sort(key=sort_fn, reverse=(sort_direction == "desc"))

    total = len(merged)

    if skip is not None and limit is not None:
        page = merged[skip:skip + limit]
    else:
        page = merged

    return total, page, sku_cost_map


# ── List (ungrouped, kept for backward compat) ──

@router.get("/")
async def list_products(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=500),
    search: Optional[str] = None,
    store_uid: Optional[str] = None,
    state: Optional[str] = None,
    has_stock: Optional[bool] = None,
    sort_field: str = "title_1",
    sort_direction: str = "asc",
    db: AsyncSession = Depends(get_db),
):
    """List products without grouping (backward compatible)."""
    query = select(Product)
    count_query = select(func.count(Product.id))

    if search:
        sp = f"%{search}%"
        sf = or_(
            Product.title_1.ilike(sp), Product.title_2.ilike(sp),
            Product.sku.ilike(sp), Product.barcode.ilike(sp),
        )
        query = query.where(sf)
        count_query = count_query.where(sf)
    if state:
        query = query.where(Product.state == state)
        count_query = count_query.where(Product.state == state)
    if has_stock is True:
        query = query.where(Product.stock_available > 0)
        count_query = count_query.where(Product.stock_available > 0)
    elif has_stock is False:
        query = query.where(Product.stock_available == 0)
        count_query = count_query.where(Product.stock_available == 0)
    if store_uid:
        query = query.where(cast(Product.store_uids, String).contains(store_uid))
        count_query = count_query.where(cast(Product.store_uids, String).contains(store_uid))

    scm = {
        "title_1": Product.title_1, "sku": Product.sku,
        "stock_available": Product.stock_available, "synced_at": Product.synced_at,
    }
    sc = scm.get(sort_field, Product.title_1)
    query = query.order_by(sc.desc() if sort_direction == "desc" else sc.asc())

    total = (await db.execute(count_query)).scalar() or 0
    result = await db.execute(query.offset(skip).limit(limit))
    products = result.scalars().all()

    suids = set()
    for p in products:
        if p.store_uids:
            suids.update(p.store_uids)
    sn = {}
    if suids:
        sr = await db.execute(select(Store.uid, Store.name).where(Store.uid.in_(list(suids))))
        sn = {r[0]: r[1] for r in sr.all()}

    return {
        "total": total, "skip": skip, "limit": limit,
        "products": [{
            "id": p.id, "uid": p.uid, "barcode": p.barcode,
            "title_1": p.title_1, "title_2": p.title_2, "sku": p.sku,
            "state": p.state, "images": p.images, "store_uids": p.store_uids,
            "stores": [{"uid": u, "name": sn.get(u, u)} for u in (p.store_uids or [])],
            "stock_available": p.stock_available, "stock_committed": p.stock_committed,
            "stock_incoming": p.stock_incoming, "exclude_from_stock": p.exclude_from_stock,
            "synced_at": p.synced_at, "frisbo_updated_at": p.frisbo_updated_at,
        } for p in products],
    }


# ── Stats ──

@router.get("/stats")
async def get_product_stats(db: AsyncSession = Depends(get_db)):
    """Product KPI statistics. Stock totals exclude flagged products."""
    total = (await db.execute(select(func.count(Product.id)))).scalar() or 0
    active = (await db.execute(
        select(func.count(Product.id)).where(Product.state == "active")
    )).scalar() or 0
    in_stock = (await db.execute(
        select(func.count(Product.id)).where(Product.stock_available > 0)
    )).scalar() or 0
    out_of_stock = (await db.execute(
        select(func.count(Product.id)).where(Product.stock_available == 0, Product.state == "active")
    )).scalar() or 0

    ne = Product.exclude_from_stock == False
    total_stock = (await db.execute(select(func.sum(Product.stock_available)).where(ne))).scalar() or 0
    total_committed = (await db.execute(select(func.sum(Product.stock_committed)).where(ne))).scalar() or 0
    excluded_count = (await db.execute(
        select(func.count(Product.id)).where(Product.exclude_from_stock == True)
    )).scalar() or 0

    return {
        "total_products": total, "active_products": active,
        "in_stock": in_stock, "out_of_stock": out_of_stock,
        "total_stock_available": total_stock, "total_stock_committed": total_committed,
        "excluded_count": excluded_count,
    }


# ── Grouped Products ──

@router.get("/grouped/")
async def list_grouped_products(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=500),
    search: Optional[str] = None,
    store_uid: Optional[str] = None,
    state: Optional[str] = None,
    has_stock: Optional[bool] = None,
    has_cost: Optional[str] = None,
    exclude_filter: Optional[str] = None,
    missing_barcode: Optional[str] = None,  # 'yes' | 'no' | None
    sort_field: str = "title_1",
    sort_direction: str = "asc",
    db: AsyncSession = Depends(get_db),
):
    """List products grouped by barcode/SKU with individual listings."""
    total, page, _ = await _fetch_and_group(
        db, search, store_uid, state, has_stock, has_cost,
        exclude_filter, missing_barcode, sort_field, sort_direction,
        skip=skip, limit=limit,
    )
    return {"total": total, "skip": skip, "limit": limit, "products": page}


# ── Exclusion Toggle ──

class ExcludeRequest(BaseModel):
    exclude: bool


@router.patch("/{product_uid}/exclude")
async def toggle_exclude(
    product_uid: str,
    body: ExcludeRequest,
    db: AsyncSession = Depends(get_db),
):
    """Toggle exclude_from_stock for a product and its entire barcode/SKU group."""
    result = await db.execute(select(Product).where(Product.uid == product_uid))
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    barcode = (product.barcode or "").strip()
    sku = (product.sku or "").strip()

    if barcode:
        # Toggle entire barcode group + any SKU-matched products
        await db.execute(
            update(Product).where(Product.barcode == barcode)
            .values(exclude_from_stock=body.exclude)
        )
        if sku:
            await db.execute(
                update(Product).where(Product.sku == sku)
                .values(exclude_from_stock=body.exclude)
            )
        affected = (await db.execute(
            select(func.count(Product.id)).where(
                or_(Product.barcode == barcode, Product.sku == sku) if sku
                else Product.barcode == barcode
            )
        )).scalar() or 1
    elif sku:
        await db.execute(
            update(Product).where(Product.sku == sku)
            .values(exclude_from_stock=body.exclude)
        )
        affected = (await db.execute(
            select(func.count(Product.id)).where(Product.sku == sku)
        )).scalar() or 1
    else:
        product.exclude_from_stock = body.exclude
        affected = 1

    await db.commit()
    return {"uid": product_uid, "exclude_from_stock": body.exclude, "affected_products": affected}


# ── Set Primary Listing ──

class SetPrimaryRequest(BaseModel):
    primary_uid: str  # UID of the listing to use as source of truth


@router.patch("/{product_uid}/set-primary")
async def set_primary_listing(
    product_uid: str,
    body: SetPrimaryRequest,
    db: AsyncSession = Depends(get_db),
):
    """
    Set which listing in a barcode/SKU group is the primary source
    for stock and image. Stores the preference on ALL products in the group
    so other calculations can use it.
    """
    result = await db.execute(select(Product).where(Product.uid == product_uid))
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    # Verify the target listing exists
    target = await db.execute(select(Product).where(Product.uid == body.primary_uid))
    if not target.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Target listing not found")

    barcode = (product.barcode or "").strip()
    sku = (product.sku or "").strip()

    # Build the group filter — same logic as _build_groups
    if barcode and sku:
        group_filter = or_(Product.barcode == barcode, Product.sku == sku)
    elif barcode:
        group_filter = Product.barcode == barcode
    elif sku:
        group_filter = Product.sku == sku
    else:
        group_filter = Product.uid == product_uid

    # Update all products in the group
    result = await db.execute(
        update(Product).where(group_filter)
        .values(primary_listing_uid=body.primary_uid)
    )
    affected = result.rowcount or 1

    await db.commit()
    return {
        "uid": product_uid,
        "primary_uid": body.primary_uid,
        "affected_products": affected,
    }


# ── Excel Export ──

@router.get("/export/excel")
async def export_products_excel(
    search: Optional[str] = None,
    store_uid: Optional[str] = None,
    state: Optional[str] = None,
    has_stock: Optional[bool] = None,
    has_cost: Optional[str] = None,
    exclude_filter: Optional[str] = None,
    missing_barcode: Optional[str] = None,
    sort_field: str = "title_1",
    sort_direction: str = "asc",
    db: AsyncSession = Depends(get_db),
):
    """Export all products (grouped) to Excel with current filters applied."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    total, products, sku_cost_map = await _fetch_and_group(
        db, search, store_uid, state, has_stock, has_cost,
        exclude_filter, missing_barcode, sort_field, sort_direction,
    )

    wb = Workbook()

    # ── Sheet 1: Grouped Products ──
    ws = wb.active
    ws.title = "Produse Grupate"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    headers = [
        "Nr.", "Titlu Produs", "Varianta", "SKU", "Cod de bare",
        "Magazine", "Stoc Disponibil", "Stoc Committed", "Stoc Incoming",
        "Cost/buc (RON)", "Stare", "Exclus din Stoc", "Nr. Listări",
        "Barcode Lipsă", "Ultima Sincronizare",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for idx, p in enumerate(products, 1):
        row = idx + 1
        stores_str = ", ".join(s.get("name", s.get("uid", "")) for s in (p.get("stores") or []))
        synced = str(p.get("synced_at") or "")[:19]

        values = [
            idx,
            p.get("title_1") or "",
            p.get("title_2") or "",
            p.get("sku") or "",
            p.get("barcode") or "",
            stores_str,
            p.get("stock_available", 0),
            p.get("stock_committed", 0),
            p.get("stock_incoming", 0),
            p.get("cost") or "",
            p.get("state") or "",
            "Da" if p.get("exclude_from_stock") else "Nu",
            p.get("grouped_count", 1),
            "Da" if p.get("has_missing_barcode") else "Nu",
            synced,
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            if col in (7, 8, 9, 10):
                cell.alignment = Alignment(horizontal="right")

    # Auto-width columns
    for col in range(1, len(headers) + 1):
        max_len = len(str(headers[col - 1]))
        for row in range(2, min(len(products) + 2, 102)):
            cell_val = ws.cell(row=row, column=col).value
            if cell_val:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = min(max_len + 3, 50)

    # ── Sheet 2: Individual Listings ──
    ws2 = wb.create_sheet("Listări Individuale")

    detail_headers = [
        "Nr.", "Grup (Barcode/SKU)", "UID Listing", "Titlu", "Varianta",
        "SKU", "Cod de bare", "Magazine", "Stoc Disponibil", "Stoc Committed",
        "Stoc Incoming", "Stare", "Organizație", "Barcode Lipsă", "Ultima Sincronizare",
    ]

    for col, header in enumerate(detail_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    row_num = 2
    for p in products:
        group_key = p.get("barcode") or p.get("sku") or p.get("uid", "")
        for listing in (p.get("listings") or []):
            stores_str = ", ".join(s.get("name", s.get("uid", "")) for s in (listing.get("stores") or []))
            synced = str(listing.get("synced_at") or "")[:19]

            values = [
                row_num - 1,
                group_key,
                listing.get("uid") or "",
                listing.get("title_1") or "",
                listing.get("title_2") or "",
                listing.get("sku") or "",
                listing.get("barcode") or "",
                stores_str,
                listing.get("stock_available", 0),
                listing.get("stock_committed", 0),
                listing.get("stock_incoming", 0),
                listing.get("state") or "",
                listing.get("organization_uid") or "",
                "Da" if listing.get("missing_barcode") else "Nu",
                synced,
            ]
            for col, val in enumerate(values, 1):
                cell = ws2.cell(row=row_num, column=col, value=val)
                cell.border = thin_border
            row_num += 1

    # Auto-width sheet 2
    for col in range(1, len(detail_headers) + 1):
        max_len = len(str(detail_headers[col - 1]))
        for row in range(2, min(row_num, 102)):
            cv = ws2.cell(row=row, column=col).value
            if cv:
                max_len = max(max_len, len(str(cv)))
        ws2.column_dimensions[ws2.cell(row=1, column=col).column_letter].width = min(max_len + 3, 50)

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=produse_export.xlsx"},
    )


# ── COGS Template Download ──

@router.get("/import/cogs-template")
async def download_cogs_template():
    """Download an Excel template for COGS import (SKU + COGS columns)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "COGS Import Template"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    # Headers
    for col, (header, width) in enumerate([("SKU", 25), ("COGS (RON)", 15)], 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    # Example rows
    examples = [
        ("SKU-001", 12.50),
        ("SKU-002", 8.75),
    ]
    for idx, (sku, cost) in enumerate(examples, 2):
        ws.cell(row=idx, column=1, value=sku).font = Font(color="999999", italic=True)
        ws.cell(row=idx, column=2, value=cost).font = Font(color="999999", italic=True)
        ws.cell(row=idx, column=1).border = thin_border
        ws.cell(row=idx, column=2).border = thin_border

    # Instructions
    ws.cell(row=5, column=1, value="⚠ Înlocuiți exemplele de mai sus cu datele dvs.").font = Font(color="CC0000", italic=True)
    ws.cell(row=6, column=1, value="Coloana A = SKU (text, obligatoriu)").font = Font(italic=True)
    ws.cell(row=7, column=1, value="Coloana B = Cost (COGS) în RON (număr)").font = Font(italic=True)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=cogs_template.xlsx"},
    )


# ── COGS Excel Import ──

from fastapi import UploadFile, File

@router.post("/import/cogs")
async def import_cogs_excel(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    """
    Import COGS from Excel file. Expected format:
    Column A = SKU, Column B = COGS (cost).
    Upserts into sku_costs table — updates existing, creates new entries.
    """
    from openpyxl import load_workbook
    from app.models.sku_cost import SkuCost
    from datetime import datetime

    # Validate file type
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Fișierul trebuie să fie format .xlsx")

    try:
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Nu pot citi fișierul Excel: {str(e)}")

    # Parse rows — skip header
    imported = []
    errors = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, max_col=2, values_only=True), 2):
        sku_val, cost_val = row[0] if len(row) > 0 else None, row[1] if len(row) > 1 else None

        if not sku_val:
            continue  # skip empty rows

        sku = str(sku_val).strip()
        if not sku:
            continue

        # Parse cost
        try:
            cost = float(cost_val) if cost_val is not None else 0.0
        except (ValueError, TypeError):
            errors.append(f"Rând {row_num}: cost invalid '{cost_val}' pentru SKU '{sku}'")
            continue

        imported.append({"sku": sku, "cost": cost})

    wb.close()

    if not imported:
        raise HTTPException(status_code=400, detail="Nu s-au găsit date valide în fișier.")

    # Upsert into sku_costs
    created = 0
    updated = 0

    for item in imported:
        result = await db.execute(
            select(SkuCost).where(SkuCost.sku == item["sku"])
        )
        existing = result.scalar_one_or_none()

        if existing:
            existing.cost = item["cost"]
            existing.updated_at = datetime.utcnow()
            updated += 1
        else:
            new_cost = SkuCost(
                sku=item["sku"],
                cost=item["cost"],
                currency="RON",
            )
            db.add(new_cost)
            created += 1

    await db.commit()

    return {
        "created": created,
        "updated": updated,
        "total_processed": len(imported),
        "errors": errors,
    }


# ── Single product ──
# IMPORTANT: This MUST be the LAST route — /{product_uid} is a catch-all
# that would shadow /export/excel, /import/cogs-template etc. if placed before them.

@router.get("/{product_uid}")
async def get_product(product_uid: str, db: AsyncSession = Depends(get_db)):
    """Get a single product by UID."""
    result = await db.execute(select(Product).where(Product.uid == product_uid))
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    sn = {}
    if product.store_uids:
        sr = await db.execute(select(Store.uid, Store.name).where(Store.uid.in_(product.store_uids)))
        sn = {r[0]: r[1] for r in sr.all()}

    return {
        "id": product.id, "uid": product.uid, "barcode": product.barcode,
        "title_1": product.title_1, "title_2": product.title_2, "sku": product.sku,
        "state": product.state, "images": product.images,
        "stores": [{"uid": u, "name": sn.get(u, u)} for u in (product.store_uids or [])],
        "stock_available": product.stock_available, "stock_committed": product.stock_committed,
        "stock_incoming": product.stock_incoming, "exclude_from_stock": product.exclude_from_stock,
        "synced_at": product.synced_at,
    }
