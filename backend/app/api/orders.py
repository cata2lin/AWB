"""
Orders API endpoints.
"""
from typing import List, Optional
from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_, cast, String
from sqlalchemy.orm import selectinload

from app.core.database import get_db
from app.core.timezone import date_str_to_utc_start, date_str_to_utc_end, romania_today_start_utc, to_bucharest_iso
from app.models import Order, Store
from app.models.order_awb import OrderAwb
from app.schemas import OrderResponse, OrderFilters, DashboardStats

router = APIRouter()


def _build_order_conditions(
    store_uids, is_printed, has_awb, has_tracking, min_items, max_items,
    search, fulfillment_status, shipment_status, aggregated_status,
    courier_names, has_shipping_cost, stale_courier, date_from, date_to,
    phone_search=None
):
    """Build reusable filter conditions for order queries."""
    from datetime import datetime, timedelta

    conditions = []
    if store_uids:
        conditions.append(Order.store_uid.in_(store_uids))
    if is_printed is not None:
        conditions.append(Order.is_printed == is_printed)
    if has_awb is True:
        conditions.append(Order.awb_pdf_url.isnot(None))
    elif has_awb is False:
        conditions.append(Order.awb_pdf_url.is_(None))
    if has_tracking is True:
        conditions.append(Order.tracking_number.isnot(None))
    elif has_tracking is False:
        conditions.append(Order.tracking_number.is_(None))
    if min_items is not None:
        conditions.append(Order.item_count >= min_items)
    if max_items is not None:
        conditions.append(Order.item_count <= max_items)
    if search:
        search_term = f"%{search}%"
        conditions.append(
            (Order.order_number.ilike(search_term)) |
            (Order.customer_name.ilike(search_term)) |
            (Order.tracking_number.ilike(search_term)) |
            (cast(Order.line_items, String).ilike(search_term))
        )
    if phone_search:
        phone_clean = phone_search.replace(" ", "")
        phone_term = f"%{phone_clean}%"
        conditions.append(
            func.replace(cast(Order.shipping_address, String), ' ', '').ilike(phone_term)
        )
    if fulfillment_status:
        conditions.append(Order.fulfillment_status.in_(fulfillment_status))
    if shipment_status:
        conditions.append(Order.shipment_status.in_(shipment_status))
    if aggregated_status:
        conditions.append(Order.aggregated_status.in_(aggregated_status))
    if courier_names:
        conditions.append(Order.courier_name.in_(courier_names))
    if stale_courier is True:
        stale_cutoff = datetime.utcnow() - timedelta(hours=72)
        conditions.append(Order.waiting_for_courier_since.isnot(None))
        conditions.append(Order.waiting_for_courier_since <= stale_cutoff)
    elif stale_courier is False:
        stale_cutoff = datetime.utcnow() - timedelta(hours=72)
        conditions.append(
            (Order.waiting_for_courier_since.is_(None)) |
            (Order.waiting_for_courier_since > stale_cutoff)
        )
    if has_shipping_cost is True:
        conditions.append(Order.transport_cost.isnot(None))
        conditions.append(Order.transport_cost > 0)
    elif has_shipping_cost is False:
        conditions.append((Order.transport_cost.is_(None)) | (Order.transport_cost == 0))
    if date_from:
        try:
            from_date = date_str_to_utc_start(date_from)
            conditions.append(Order.frisbo_created_at >= from_date)
        except ValueError:
            pass
    if date_to:
        try:
            to_date = date_str_to_utc_end(date_to)
            conditions.append(Order.frisbo_created_at <= to_date)
        except ValueError:
            pass
    return conditions


@router.get("", response_model=List[OrderResponse])
async def get_orders(
    store_uids: Optional[List[str]] = Query(None),
    is_printed: Optional[bool] = None,
    has_awb: Optional[bool] = None,
    has_tracking: Optional[bool] = None,
    min_items: Optional[int] = None,
    max_items: Optional[int] = None,
    search: Optional[str] = None,
    phone_search: Optional[str] = Query(None, description="Search by phone number in shipping address"),
    fulfillment_status: Optional[List[str]] = Query(None, description="Filter by fulfillment status (multi)"),
    shipment_status: Optional[List[str]] = Query(None, description="Filter by shipment status (multi)"),
    aggregated_status: Optional[List[str]] = Query(None, description="Filter by workflow/aggregated status (multi)"),
    courier_names: Optional[List[str]] = Query(None, description="Filter by courier name (multi)"),
    has_shipping_cost: Optional[bool] = Query(None, description="Filter by whether order has shipping cost"),
    stale_courier: Optional[bool] = Query(None, description="Filter orders waiting for courier > 72 hours"),
    date_from: Optional[str] = Query(None, description="Filter from date (YYYY-MM-DD)"),
    date_to: Optional[str] = Query(None, description="Filter to date (YYYY-MM-DD)"),
    sort_field: Optional[str] = Query("frisbo_created_at", description="Field to sort by"),
    sort_direction: Optional[str] = Query("desc", description="Sort direction: asc or desc"),
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=500),
    db: AsyncSession = Depends(get_db)
):
    """Get orders with optional filters and sorting."""
    from datetime import datetime, timedelta
    
    # Use selectinload to eagerly load Store relationship (required for async SQLAlchemy)
    query = select(Order).options(selectinload(Order.store))
    
    # Apply filters
    conditions = []
    
    if store_uids:
        conditions.append(Order.store_uid.in_(store_uids))
    
    if is_printed is not None:
        conditions.append(Order.is_printed == is_printed)
    
    if has_awb is True:
        conditions.append(Order.awb_pdf_url.isnot(None))
    elif has_awb is False:
        conditions.append(Order.awb_pdf_url.is_(None))
    
    if has_tracking is True:
        conditions.append(Order.tracking_number.isnot(None))
    elif has_tracking is False:
        conditions.append(Order.tracking_number.is_(None))
    
    if min_items is not None:
        conditions.append(Order.item_count >= min_items)
    
    if max_items is not None:
        conditions.append(Order.item_count <= max_items)
    
    if search:
        search_term = f"%{search}%"
        conditions.append(
            (Order.order_number.ilike(search_term)) |
            (Order.customer_name.ilike(search_term)) |
            (Order.tracking_number.ilike(search_term)) |
            (cast(Order.line_items, String).ilike(search_term))
        )
    
    if phone_search:
        phone_clean = phone_search.replace(" ", "")
        phone_term = f"%{phone_clean}%"
        conditions.append(
            func.replace(cast(Order.shipping_address, String), ' ', '').ilike(phone_term)
        )
    
    # Status filters (support multi-select)
    if fulfillment_status:
        conditions.append(Order.fulfillment_status.in_(fulfillment_status))
    
    if shipment_status:
        conditions.append(Order.shipment_status.in_(shipment_status))
    
    if aggregated_status:
        conditions.append(Order.aggregated_status.in_(aggregated_status))
    
    # Courier filter (multi-select)
    if courier_names:
        conditions.append(Order.courier_name.in_(courier_names))
    
    # Stale courier filter (waiting for courier > 72 hours)
    if stale_courier is True:
        stale_cutoff = datetime.utcnow() - timedelta(hours=72)
        conditions.append(Order.waiting_for_courier_since.isnot(None))
        conditions.append(Order.waiting_for_courier_since <= stale_cutoff)
    elif stale_courier is False:
        # Orders NOT stale (either not waiting, or waiting < 72h)
        stale_cutoff = datetime.utcnow() - timedelta(hours=72)
        conditions.append(
            (Order.waiting_for_courier_since.is_(None)) |
            (Order.waiting_for_courier_since > stale_cutoff)
        )
    
    # Shipping cost filter
    if has_shipping_cost is True:
        conditions.append(Order.transport_cost.isnot(None))
        conditions.append(Order.transport_cost > 0)
    elif has_shipping_cost is False:
        conditions.append((Order.transport_cost.is_(None)) | (Order.transport_cost == 0))
    
    if date_from:
        try:
            from_date = date_str_to_utc_start(date_from)
            conditions.append(Order.frisbo_created_at >= from_date)
        except ValueError:
            pass
    
    if date_to:
        try:
            to_date = date_str_to_utc_end(date_to)
            conditions.append(Order.frisbo_created_at <= to_date)
        except ValueError:
            pass
    
    if conditions:
        query = query.where(and_(*conditions))
    
    # Server-side sorting - map field names to Order model columns
    sort_column_map = {
        "frisbo_created_at": Order.frisbo_created_at,
        "order_number": Order.order_number,
        "customer_name": Order.customer_name,
        "item_count": Order.item_count,
        "tracking_number": Order.tracking_number,
        "courier_name": Order.courier_name,
        "transport_cost": Order.transport_cost,
        "total_price": Order.total_price,
        "fulfilled_at": Order.fulfilled_at,
        "synced_at": Order.synced_at,
        "store_name": Store.name,  # Joined column from Store relationship
    }
    
    # For store_name sorting, ensure we join the Store table explicitly
    sort_col = sort_column_map.get(sort_field, Order.frisbo_created_at)
    if sort_field == "store_name":
        query = query.join(Store, Order.store_uid == Store.uid, isouter=True)
    
    if sort_direction == "asc":
        query = query.order_by(sort_col.asc().nulls_last())
    else:
        query = query.order_by(sort_col.desc().nulls_last())
    
    query = query.offset(skip).limit(limit)
    
    result = await db.execute(query)
    orders = result.scalars().all()
    
    # Enrich with store info
    response = []
    for order in orders:
        order_dict = {
            "id": order.id,
            "uid": order.uid,
            "order_number": order.order_number or "",
            "store_uid": order.store_uid or "",
            "customer_name": order.customer_name or "Unknown",
            "customer_email": order.customer_email,
            "shipping_address": order.shipping_address if order.shipping_address else {},
            "line_items": order.line_items if order.line_items else [],
            "item_count": order.item_count or 0,
            "unique_sku_count": order.unique_sku_count or 0,
            "tracking_number": order.tracking_number,
            "courier_name": order.courier_name,
            "awb_pdf_url": order.awb_pdf_url,
            "fulfillment_status": order.fulfillment_status or "unknown",
            "shipment_status": order.shipment_status,
            "aggregated_status": order.aggregated_status,
            "is_printed": order.is_printed,
            "frisbo_created_at": order.frisbo_created_at,
            "fulfilled_at": order.fulfilled_at,
            "synced_at": order.synced_at,
            "printed_at": order.printed_at,
            "store_name": order.store.name if order.store else None,
            "store_color": order.store.color_code if order.store else "#6366f1",
            # Multi-AWB
            "awb_count": order.awb_count or 1,
            "awb_count_manual": order.awb_count_manual or False,
            # Shipping data
            "package_count": order.package_count,
            "package_weight": order.package_weight,
            "transport_cost": order.transport_cost,
            "shipping_data_source": order.shipping_data_source,
            "shipping_data_manual": order.shipping_data_manual or False,
            # Financial
            "total_price": order.total_price,
            "subtotal_price": order.subtotal_price,
            "currency": order.currency,
            # Waiting for courier data
            "waiting_for_courier_since": order.waiting_for_courier_since,
            "is_stale_courier": (
                order.waiting_for_courier_since is not None and
                (datetime.utcnow() - order.waiting_for_courier_since).total_seconds() > 72 * 3600
            ),
        }
        # We'll add awb_count_actual later via a subquery if needed
        response.append(OrderResponse(**order_dict))
    
    return response


@router.get("/couriers")
async def get_couriers(db: AsyncSession = Depends(get_db)):
    """Get distinct courier names for filter dropdown."""
    result = await db.execute(
        select(Order.courier_name)
        .where(Order.courier_name.isnot(None))
        .distinct()
        .order_by(Order.courier_name)
    )
    couriers = [row[0] for row in result.fetchall() if row[0]]
    return {"couriers": couriers}


@router.get("/filter-options")
async def get_filter_options(db: AsyncSession = Depends(get_db)):
    """Get all unique filter option values from the database."""
    # Get unique shipment statuses
    shipment_result = await db.execute(
        select(Order.shipment_status)
        .where(Order.shipment_status.isnot(None))
        .distinct()
    )
    shipment_statuses = sorted([row[0] for row in shipment_result.fetchall() if row[0]])
    
    # Get unique fulfillment statuses
    fulfillment_result = await db.execute(
        select(Order.fulfillment_status)
        .where(Order.fulfillment_status.isnot(None))
        .distinct()
    )
    fulfillment_statuses = sorted([row[0] for row in fulfillment_result.fetchall() if row[0]])
    
    # Get unique aggregated/workflow statuses
    workflow_result = await db.execute(
        select(Order.aggregated_status)
        .where(Order.aggregated_status.isnot(None))
        .distinct()
    )
    workflow_statuses = sorted([row[0] for row in workflow_result.fetchall() if row[0]])
    
    # Get unique couriers
    courier_result = await db.execute(
        select(Order.courier_name)
        .where(Order.courier_name.isnot(None))
        .distinct()
    )
    couriers = sorted([row[0] for row in courier_result.fetchall() if row[0]])
    
    # Get count of orders with tracking
    tracking_count_result = await db.execute(
        select(func.count(Order.id)).where(Order.tracking_number.isnot(None))
    )
    orders_with_tracking = tracking_count_result.scalar() or 0
    
    return {
        "shipment_statuses": shipment_statuses,
        "fulfillment_statuses": fulfillment_statuses,
        "workflow_statuses": workflow_statuses,
        "couriers": couriers,
        "orders_with_tracking": orders_with_tracking
    }


@router.get("/count")
async def get_order_count(
    store_uids: Optional[List[str]] = Query(None),
    is_printed: Optional[bool] = None,
    has_awb: Optional[bool] = None,
    has_tracking: Optional[bool] = None,
    min_items: Optional[int] = None,
    max_items: Optional[int] = None,
    search: Optional[str] = None,
    phone_search: Optional[str] = Query(None, description="Search by phone number in shipping address"),
    fulfillment_status: Optional[List[str]] = Query(None),
    shipment_status: Optional[List[str]] = Query(None),
    aggregated_status: Optional[List[str]] = Query(None),
    courier_names: Optional[List[str]] = Query(None),
    has_shipping_cost: Optional[bool] = Query(None),
    stale_courier: Optional[bool] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db)
):
    """Get order counts with same filters as main orders endpoint."""
    from datetime import datetime, timedelta
    
    query = select(func.count(Order.id))
    
    conditions = []
    if store_uids:
        conditions.append(Order.store_uid.in_(store_uids))
    if is_printed is not None:
        conditions.append(Order.is_printed == is_printed)
    if has_awb is True:
        conditions.append(Order.awb_pdf_url.isnot(None))
    elif has_awb is False:
        conditions.append(Order.awb_pdf_url.is_(None))
    if has_tracking is True:
        conditions.append(Order.tracking_number.isnot(None))
    elif has_tracking is False:
        conditions.append(Order.tracking_number.is_(None))
    if min_items is not None:
        conditions.append(Order.item_count >= min_items)
    if max_items is not None:
        conditions.append(Order.item_count <= max_items)
    if search:
        search_term = f"%{search}%"
        conditions.append(
            (Order.order_number.ilike(search_term)) |
            (Order.customer_name.ilike(search_term)) |
            (Order.tracking_number.ilike(search_term)) |
            (cast(Order.line_items, String).ilike(search_term))
        )
    if phone_search:
        phone_clean = phone_search.replace(" ", "")
        phone_term = f"%{phone_clean}%"
        conditions.append(
            func.replace(cast(Order.shipping_address, String), ' ', '').ilike(phone_term)
        )
    if fulfillment_status:
        conditions.append(Order.fulfillment_status.in_(fulfillment_status))
    if shipment_status:
        conditions.append(Order.shipment_status.in_(shipment_status))
    if aggregated_status:
        conditions.append(Order.aggregated_status.in_(aggregated_status))
    if courier_names:
        conditions.append(Order.courier_name.in_(courier_names))
    if has_shipping_cost is True:
        conditions.append(Order.transport_cost.isnot(None))
        conditions.append(Order.transport_cost > 0)
    elif has_shipping_cost is False:
        conditions.append((Order.transport_cost.is_(None)) | (Order.transport_cost == 0))
    if stale_courier is True:
        stale_cutoff = datetime.utcnow() - timedelta(hours=72)
        conditions.append(Order.waiting_for_courier_since.isnot(None))
        conditions.append(Order.waiting_for_courier_since <= stale_cutoff)
    elif stale_courier is False:
        stale_cutoff = datetime.utcnow() - timedelta(hours=72)
        conditions.append(
            (Order.waiting_for_courier_since.is_(None)) |
            (Order.waiting_for_courier_since > stale_cutoff)
        )
    if date_from:
        try:
            from_date = date_str_to_utc_start(date_from)
            conditions.append(Order.frisbo_created_at >= from_date)
        except ValueError:
            pass
    if date_to:
        try:
            to_date = date_str_to_utc_end(date_to)
            conditions.append(Order.frisbo_created_at <= to_date)
        except ValueError:
            pass
    
    if conditions:
        query = query.where(and_(*conditions))
    
    result = await db.execute(query)
    count = result.scalar()
    
    return {"count": count}


@router.get("/totals")
async def get_order_totals(
    store_uids: Optional[List[str]] = Query(None),
    is_printed: Optional[bool] = None,
    has_awb: Optional[bool] = None,
    has_tracking: Optional[bool] = None,
    min_items: Optional[int] = None,
    max_items: Optional[int] = None,
    search: Optional[str] = None,
    phone_search: Optional[str] = Query(None, description="Search by phone number in shipping address"),
    fulfillment_status: Optional[List[str]] = Query(None),
    shipment_status: Optional[List[str]] = Query(None),
    aggregated_status: Optional[List[str]] = Query(None),
    courier_names: Optional[List[str]] = Query(None),
    has_shipping_cost: Optional[bool] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db)
):
    """Get total order value in RON (with currency conversion) for filtered orders."""
    from datetime import datetime
    from app.models.exchange_rate import ExchangeRate
    
    # Build same conditions as main query
    conditions = []
    if store_uids:
        conditions.append(Order.store_uid.in_(store_uids))
    if is_printed is not None:
        conditions.append(Order.is_printed == is_printed)
    if has_awb is True:
        conditions.append(Order.awb_pdf_url.isnot(None))
    elif has_awb is False:
        conditions.append(Order.awb_pdf_url.is_(None))
    if has_tracking is True:
        conditions.append(Order.tracking_number.isnot(None))
    elif has_tracking is False:
        conditions.append(Order.tracking_number.is_(None))
    if min_items is not None:
        conditions.append(Order.item_count >= min_items)
    if max_items is not None:
        conditions.append(Order.item_count <= max_items)
    if search:
        search_term = f"%{search}%"
        conditions.append(
            (Order.order_number.ilike(search_term)) |
            (Order.customer_name.ilike(search_term)) |
            (Order.tracking_number.ilike(search_term)) |
            (cast(Order.line_items, String).ilike(search_term))
        )
    if phone_search:
        phone_clean = phone_search.replace(" ", "")
        phone_term = f"%{phone_clean}%"
        conditions.append(
            func.replace(cast(Order.shipping_address, String), ' ', '').ilike(phone_term)
        )
    if fulfillment_status:
        conditions.append(Order.fulfillment_status.in_(fulfillment_status))
    if shipment_status:
        conditions.append(Order.shipment_status.in_(shipment_status))
    if aggregated_status:
        conditions.append(Order.aggregated_status.in_(aggregated_status))
    if courier_names:
        conditions.append(Order.courier_name.in_(courier_names))
    if has_shipping_cost is True:
        conditions.append(Order.transport_cost.isnot(None))
        conditions.append(Order.transport_cost > 0)
    elif has_shipping_cost is False:
        conditions.append((Order.transport_cost.is_(None)) | (Order.transport_cost == 0))
    if date_from:
        try:
            from_date = date_str_to_utc_start(date_from)
            conditions.append(Order.frisbo_created_at >= from_date)
        except ValueError:
            pass
    if date_to:
        try:
            to_date = date_str_to_utc_end(date_to)
            conditions.append(Order.frisbo_created_at <= to_date)
        except ValueError:
            pass
    
    # Aggregate total_price grouped by currency
    currency_col = func.coalesce(Order.currency, 'RON').label('currency')
    query = select(
        currency_col,
        func.sum(Order.total_price).label('total'),
        func.count(Order.id).label('count')
    ).where(Order.total_price.isnot(None))
    
    if conditions:
        query = query.where(and_(*conditions))
    
    query = query.group_by(currency_col)
    
    result = await db.execute(query)
    rows = result.all()
    
    # Get latest exchange rates for non-RON currencies
    rate_map = {'RON': 1.0}
    non_ron_currencies = [r.currency for r in rows if r.currency != 'RON']
    if non_ron_currencies:
        for curr in non_ron_currencies:
            rate_result = await db.execute(
                select(ExchangeRate)
                .where(ExchangeRate.currency == curr)
                .order_by(ExchangeRate.rate_date.desc())
                .limit(1)
            )
            rate = rate_result.scalar_one_or_none()
            if rate:
                rate_map[curr] = rate.rate / rate.multiplier
            else:
                rate_map[curr] = 1.0  # Fallback if no rate found
    
    # Calculate total in RON
    total_ron = 0.0
    per_currency = []
    total_count = 0
    for row in rows:
        amount_ron = (row.total or 0) * rate_map.get(row.currency, 1.0)
        total_ron += amount_ron
        total_count += row.count
        per_currency.append({
            'currency': row.currency,
            'total': round(row.total or 0, 2),
            'count': row.count,
            'rate_to_ron': round(rate_map.get(row.currency, 1.0), 4),
            'total_ron': round(amount_ron, 2),
        })
    
    return {
        'total_ron': round(total_ron, 2),
        'total_count': total_count,
        'per_currency': per_currency,
    }

@router.get("/stats", response_model=DashboardStats)
async def get_dashboard_stats(db: AsyncSession = Depends(get_db)):
    """Get dashboard statistics."""
    from datetime import datetime, timedelta
    
    today_start = romania_today_start_utc()
    
    # Total orders
    total_result = await db.execute(select(func.count(Order.id)))
    total_orders = total_result.scalar() or 0
    
    # Unprinted orders
    unprinted_result = await db.execute(
        select(func.count(Order.id)).where(Order.is_printed == False)
    )
    unprinted_orders = unprinted_result.scalar() or 0
    
    # Total stores
    stores_result = await db.execute(select(func.count(Store.id)))
    total_stores = stores_result.scalar() or 0
    
    # Active rules
    from app.models import Rule
    rules_result = await db.execute(
        select(func.count(Rule.id)).where(Rule.is_active == True)
    )
    active_rules = rules_result.scalar() or 0
    
    # Batches today
    from app.models import PrintBatch
    batches_result = await db.execute(
        select(func.count(PrintBatch.id)).where(PrintBatch.created_at >= today_start)
    )
    batches_today = batches_result.scalar() or 0
    
    # Orders printed today
    printed_today_result = await db.execute(
        select(func.count(Order.id)).where(
            and_(Order.is_printed == True, Order.printed_at >= today_start)
        )
    )
    orders_printed_today = printed_today_result.scalar() or 0
    
    # Stale courier count (waiting > 72h)
    stale_cutoff = today_start - timedelta(hours=72)
    stale_result = await db.execute(
        select(func.count(Order.id)).where(
            Order.waiting_for_courier_since.isnot(None),
            Order.waiting_for_courier_since <= datetime.utcnow() - timedelta(hours=72),
        )
    )
    stale_courier_count = stale_result.scalar() or 0
    
    return DashboardStats(
        total_orders=total_orders,
        unprinted_orders=unprinted_orders,
        total_stores=total_stores,
        active_rules=active_rules,
        batches_today=batches_today,
        orders_printed_today=orders_printed_today,
        stale_courier_count=stale_courier_count,
    )


@router.post("/mark-all-printed")
async def mark_all_orders_printed(db: AsyncSession = Depends(get_db)):
    """Mark all orders in the database as printed."""
    from sqlalchemy import update
    from datetime import datetime
    
    result = await db.execute(
        update(Order)
        .where(Order.is_printed == False)
        .values(is_printed=True, printed_at=datetime.utcnow())
    )
    await db.commit()
    
    return {"message": f"Marked {result.rowcount} orders as printed"}


@router.get("/export")
async def export_orders_excel(
    store_uids: Optional[List[str]] = Query(None),
    is_printed: Optional[bool] = None,
    has_awb: Optional[bool] = None,
    has_tracking: Optional[bool] = None,
    min_items: Optional[int] = None,
    max_items: Optional[int] = None,
    search: Optional[str] = None,
    phone_search: Optional[str] = Query(None, description="Search by phone number in shipping address"),
    fulfillment_status: Optional[List[str]] = Query(None),
    shipment_status: Optional[List[str]] = Query(None),
    aggregated_status: Optional[List[str]] = Query(None),
    courier_names: Optional[List[str]] = Query(None),
    has_shipping_cost: Optional[bool] = Query(None),
    stale_courier: Optional[bool] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    sort_field: Optional[str] = Query("frisbo_created_at"),
    sort_direction: Optional[str] = Query("desc"),
    db: AsyncSession = Depends(get_db)
):
    """Export filtered orders as an Excel (.xlsx) file.

    Applies the same filters as the main orders endpoint but returns ALL
    matching rows (no pagination) as a downloadable spreadsheet.
    """
    import io
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # ── Build query with same filters ──
    query = select(Order).options(selectinload(Order.store))
    conditions = _build_order_conditions(
        store_uids, is_printed, has_awb, has_tracking, min_items, max_items,
        search, fulfillment_status, shipment_status, aggregated_status,
        courier_names, has_shipping_cost, stale_courier, date_from, date_to,
        phone_search=phone_search
    )
    if conditions:
        query = query.where(and_(*conditions))

    # Sorting
    sort_column_map = {
        "frisbo_created_at": Order.frisbo_created_at,
        "order_number": Order.order_number,
        "customer_name": Order.customer_name,
        "item_count": Order.item_count,
        "tracking_number": Order.tracking_number,
        "courier_name": Order.courier_name,
        "transport_cost": Order.transport_cost,
        "total_price": Order.total_price,
        "fulfilled_at": Order.fulfilled_at,
        "store_name": Store.name,
    }
    sort_col = sort_column_map.get(sort_field, Order.frisbo_created_at)
    if sort_field == "store_name":
        query = query.join(Store, Order.store_uid == Store.uid, isouter=True)
    if sort_direction == "asc":
        query = query.order_by(sort_col.asc().nulls_last())
    else:
        query = query.order_by(sort_col.desc().nulls_last())

    result = await db.execute(query)
    orders = result.scalars().all()

    # ── Build Excel workbook ──
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"

    # Column definitions matching the Orders table view
    headers = [
        "Order #",
        "Customer",
        "Store",
        "Items",
        "Total",
        "Currency",
        "Fulfillment Status",
        "Shipment Status",
        "Workflow Status",
        "Courier",
        "Tracking #",
        "Transport Cost (RON)",
        "Cost Source",
        "Created Date",
        "Fulfilled Date",
        "Printed",
    ]

    # Styling
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        bottom=Side(style="thin", color="D4D4D8")
    )
    date_format = "YYYY-MM-DD HH:MM"

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Write data rows
    for row_idx, order in enumerate(orders, 2):
        store_name = order.store.name if order.store else ""
        line_items_text = ""
        if order.line_items:
            items = []
            for li in order.line_items:
                sku = li.get("inventory_item", {}).get("sku", "N/A") if isinstance(li, dict) else "N/A"
                qty = li.get("quantity", 1) if isinstance(li, dict) else 1
                items.append(f"{sku} x{qty}")
            line_items_text = ", ".join(items)

        row_data = [
            order.order_number or "",
            order.customer_name or "",
            store_name,
            order.item_count or 0,
            float(order.total_price) if order.total_price is not None else None,
            order.currency or "RON",
            (order.fulfillment_status or "").replace("_", " ").title(),
            (order.shipment_status or "").replace("_", " ").title(),
            (order.aggregated_status or "").replace("_", " ").title(),
            order.courier_name or "",
            order.tracking_number or "",
            float(order.transport_cost) if order.transport_cost is not None else None,
            order.shipping_data_source or "",
            order.frisbo_created_at,
            order.fulfilled_at,
            "Yes" if order.is_printed else "No",
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            # Format date columns
            if col_idx in (14, 15) and value is not None:
                cell.number_format = date_format

    # Auto-fit column widths (approximate)
    col_widths = [14, 22, 18, 8, 12, 10, 18, 18, 18, 16, 22, 18, 14, 18, 18, 10]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(headers)).column_letter}{len(orders) + 1}"

    # ── Stream response ──
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Build descriptive filename
    now = datetime.utcnow().strftime("%Y%m%d_%H%M")
    filename = f"orders_export_{now}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# ────────────────────────────────────────────────────────────────────────────
# Cross-store duplicate order detection
# ────────────────────────────────────────────────────────────────────────────
# Implements the logic from two Shopify Flows:
#   1. "Tag duplicate orders with identical products" — same customer, same
#      SKU set within 7 days → "product_duplicate"
#   2. "Comanda multipla cu sapte zile in urma" — same customer, >1 order
#      within 7 days → "repeat_customer"
#
# Our version is CROSS-STORE: it matches by customer_email across all stores.
# ────────────────────────────────────────────────────────────────────────────

def _sku_fingerprint(line_items) -> str:
    """Create a sorted, canonical fingerprint from an order's SKUs + quantities.

    Why include quantity: 2× "lavete" is different from 1× "lavete".
    Matching Shopify Flow 1's variant-ID comparison logic.
    """
    if not line_items or not isinstance(line_items, list):
        return ""
    parts = []
    for item in line_items:
        inv = item.get("inventory_item", {}) or {}
        sku = inv.get("sku", "") or ""
        qty = item.get("quantity", 1) or 1
        if sku:
            parts.append(f"{sku}:{qty}")
    return ",".join(sorted(parts))


@router.get("/duplicates")
async def get_duplicate_orders(
    date_from: Optional[str] = Query(None, description="Filter from date (YYYY-MM-DD)"),
    date_to: Optional[str] = Query(None, description="Filter to date (YYYY-MM-DD)"),
    store_uids: Optional[List[str]] = Query(None, description="Limit to specific stores"),
    window_days: int = Query(7, ge=1, le=90, description="Duplicate window in days"),
    dup_type: Optional[str] = Query(None, description="Filter: 'product_duplicate', 'repeat_customer', or None for all"),
    search: Optional[str] = Query(None, description="Search by customer name, email, or order number"),
    phone_search: Optional[str] = Query(None, description="Search by phone number (normalized)"),
    sort_by: Optional[str] = Query("date", description="Sort groups by: 'date', 'orders', 'type'"),
    sort_dir: Optional[str] = Query("desc", description="Sort direction: 'asc' or 'desc'"),
    page: int = Query(0, ge=0),
    limit: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db)
):
    """Detect cross-store duplicate orders.

    Groups orders by customer email. Within each group, orders placed
    within ``window_days`` of each other are flagged as duplicates.

    Defaults to last 14 days if no date range is specified for performance.

    Returns paginated duplicate groups, each containing all matching orders
    with full details (store, products, status, timestamps).
    """
    import re
    from datetime import datetime, timedelta
    from collections import defaultdict

    def _normalize_phone(raw: str) -> str:
        """Strip spaces/dashes and unify country prefixes to bare digits."""
        if not raw:
            return ""
        d = re.sub(r'[\s\-().]+', '', raw)
        if d.startswith('+40'):
            d = d[3:]
        elif d.startswith('+4') and len(d) > 10:
            d = d[2:]
        elif d.startswith('040'):
            d = d[3:]
        elif d.startswith('40') and len(d) > 9:
            d = d[2:]
        elif d.startswith('04'):
            d = d[2:]
        elif d.startswith('+'):
            d = d[1:]  # Other country codes: strip +
        return d

    # ── Default to last 14 days if no dates provided ──
    if not date_from and not date_to:
        default_start = datetime.utcnow() - timedelta(days=14)
        date_from_dt = default_start
    else:
        date_from_dt = None

    # ── Build date filter ──
    conditions = []
    if date_from:
        try:
            conditions.append(Order.frisbo_created_at >= date_str_to_utc_start(date_from))
        except ValueError:
            pass
    elif date_from_dt:
        conditions.append(Order.frisbo_created_at >= date_from_dt)

    if date_to:
        try:
            conditions.append(Order.frisbo_created_at <= date_str_to_utc_end(date_to))
        except ValueError:
            pass
    if store_uids:
        conditions.append(Order.store_uid.in_(store_uids))

    # Only consider orders with a customer email (cross-store key)
    conditions.append(Order.customer_email.isnot(None))
    conditions.append(Order.customer_email != "")

    # ── Step 1: Find emails with >1 order in the period (uses ix_orders_email_date) ──
    email_subq = (
        select(Order.customer_email)
        .where(and_(*conditions))
        .group_by(Order.customer_email)
        .having(func.count(Order.id) > 1)
    ).subquery()

    # ── Step 2: Fetch only orders belonging to those emails ──
    query = (
        select(Order)
        .options(selectinload(Order.store))
        .where(
            and_(
                *conditions,
                Order.customer_email.in_(select(email_subq.c.customer_email))
            )
        )
        .order_by(Order.customer_email, Order.frisbo_created_at.asc())
    )

    result = await db.execute(query)
    all_orders = result.scalars().all()

    # ── Group by email ──
    email_groups: dict = defaultdict(list)
    for order in all_orders:
        email_groups[order.customer_email.lower().strip()].append(order)

    duplicate_groups = []

    for email, orders_list in email_groups.items():
        if len(orders_list) < 2:
            continue

        # Orders already sorted by date from SQL — linear single-pass clustering
        # Merge orders within window_days of the cluster's latest timestamp
        window_secs = window_days * 86400
        clusters = []
        current_cluster = [0]
        cluster_end = (orders_list[0].frisbo_created_at or datetime.min)

        for i in range(1, len(orders_list)):
            dt_i = orders_list[i].frisbo_created_at or datetime.min
            if (dt_i - cluster_end).total_seconds() <= window_secs:
                current_cluster.append(i)
                if dt_i > cluster_end:
                    cluster_end = dt_i
            else:
                if len(current_cluster) >= 2:
                    clusters.append(current_cluster)
                current_cluster = [i]
                cluster_end = dt_i

        if len(current_cluster) >= 2:
            clusters.append(current_cluster)

        for cluster_indices in clusters:
            cluster_orders = [orders_list[idx] for idx in cluster_indices]

            # Classify duplicate type
            fingerprints = [_sku_fingerprint(o.line_items) for o in cluster_orders]
            has_product_dup = len(fingerprints) != len(set(fingerprints)) and all(f for f in fingerprints)

            group_type = "product_duplicate" if has_product_dup else "repeat_customer"

            # Apply type filter if requested
            if dup_type and group_type != dup_type:
                continue

            # Collect unique stores
            store_names = set()
            for o in cluster_orders:
                if o.store:
                    store_names.add(o.store.name)

            # Build order summaries
            order_summaries = []
            for o in cluster_orders:
                items_raw = o.line_items or []
                skus = []
                for item in (items_raw if isinstance(items_raw, list) else []):
                    inv = item.get("inventory_item", {}) or {}
                    sku = inv.get("sku", "") or ""
                    title = inv.get("title_1", "") or ""
                    qty = item.get("quantity", 1) or 1
                    price = item.get("price", 0) or 0
                    if sku:
                        skus.append({
                            "sku": sku,
                            "title": title,
                            "quantity": qty,
                            "price": float(price),
                        })

                addr = o.shipping_address or {}
                shopify_domain = o.store.shopify_domain if o.store else None
                # Build Shopify admin search URL using the full order number
                shopify_order_url = None
                if shopify_domain and o.order_number:
                    from urllib.parse import quote
                    shopify_order_url = f"https://{shopify_domain}/admin/orders?query={quote(o.order_number)}"
                order_summaries.append({
                    "uid": o.uid,
                    "order_number": o.order_number or "",
                    "store_name": o.store.name if o.store else "",
                    "store_color": o.store.color_code if o.store else "#6366f1",
                    "store_uid": o.store_uid or "",
                    "shopify_domain": shopify_domain or "",
                    "shopify_order_url": shopify_order_url or "",
                    "customer_name": o.customer_name or "",
                    "customer_email": o.customer_email or "",
                    "phone": addr.get("phone", ""),
                    "city": addr.get("city", ""),
                    "country": addr.get("country_code", ""),
                    "address": f"{addr.get('address1', '')} {addr.get('address2', '')}".strip(),
                    "total_price": o.total_price,
                    "subtotal_price": o.subtotal_price,
                    "total_discounts": o.total_discounts,
                    "currency": o.currency or "RON",
                    "item_count": o.item_count or 0,
                    "items": skus,
                    "fulfillment_status": o.fulfillment_status,
                    "shipment_status": o.shipment_status,
                    "aggregated_status": o.aggregated_status,
                    "financial_status": o.financial_status,
                    "payment_gateway": o.payment_gateway,
                    "tracking_number": o.tracking_number,
                    "courier_name": o.courier_name,
                    "frisbo_created_at": to_bucharest_iso(o.frisbo_created_at),
                    "sku_fingerprint": _sku_fingerprint(o.line_items),
                })

            latest_date = max(
                (o.frisbo_created_at for o in cluster_orders if o.frisbo_created_at),
                default=datetime.min,
            )

            duplicate_groups.append({
                "customer_email": email,
                "customer_name": cluster_orders[0].customer_name or "",
                "group_type": group_type,
                "order_count": len(cluster_orders),
                "stores": sorted(store_names),
                "is_cross_store": len(store_names) > 1,
                "date_range": {
                    "earliest": to_bucharest_iso(cluster_orders[0].frisbo_created_at),
                    "latest": to_bucharest_iso(cluster_orders[-1].frisbo_created_at),
                },
                "_latest_ts": latest_date,
                "orders": order_summaries,
            })

    # ── Server-side search filter (applied AFTER clustering, BEFORE pagination) ──
    if search:
        q = search.lower().strip()
        duplicate_groups = [
            g for g in duplicate_groups
            if q in g["customer_email"].lower()
            or q in g["customer_name"].lower()
            or any(q in o["order_number"].lower() for o in g["orders"])
        ]

    if phone_search:
        norm_q = _normalize_phone(phone_search)
        if norm_q:
            duplicate_groups = [
                g for g in duplicate_groups
                if any(
                    norm_q in _normalize_phone(o.get("phone", ""))
                    for o in g["orders"]
                )
            ]

    # ── Sorting ──
    reverse = sort_dir != "asc"

    if sort_by == "orders":
        duplicate_groups.sort(key=lambda g: g["order_count"], reverse=reverse)
    elif sort_by == "type":
        duplicate_groups.sort(key=lambda g: (
            g["group_type"] == "product_duplicate",
            g["is_cross_store"],
        ), reverse=reverse)
    else:  # default: "date"
        duplicate_groups.sort(key=lambda g: g["_latest_ts"], reverse=reverse)

    total_groups = len(duplicate_groups)
    total_orders_in_groups = sum(g["order_count"] for g in duplicate_groups)

    # Paginate and strip internal fields
    paginated = duplicate_groups[page * limit : (page + 1) * limit]
    for g in paginated:
        g.pop("_latest_ts", None)

    return {
        "total_groups": total_groups,
        "total_duplicate_orders": total_orders_in_groups,
        "page": page,
        "limit": limit,
        "groups": paginated,
    }


@router.get("/{order_uid}", response_model=OrderResponse)
async def get_order(order_uid: str, db: AsyncSession = Depends(get_db)):
    """Get a specific order by UID."""
    result = await db.execute(
        select(Order).where(Order.uid == order_uid)
    )
    order = result.scalar_one_or_none()
    
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    return order


@router.get("/{order_uid}/awbs")
async def get_order_awbs(order_uid: str, db: AsyncSession = Depends(get_db)):
    """Get all AWB records for an order, including billable status."""
    from app.models.order_awb import is_billable_status
    
    # Find order
    result = await db.execute(select(Order).where(Order.uid == order_uid))
    order = result.scalar_one_or_none()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    # Get all AWBs
    awb_result = await db.execute(
        select(OrderAwb)
        .where(OrderAwb.order_id == order.id)
        .order_by(OrderAwb.awb_type.asc(), OrderAwb.created_at.desc())
    )
    awbs = awb_result.scalars().all()
    
    # Calculate billable totals
    billable_cost = 0
    billable_count = 0
    for awb in awbs:
        if awb.awb_type == 'outbound' and awb.transport_cost and is_billable_status(awb.csv_status):
            billable_cost += awb.transport_cost
            billable_count += 1
    
    return {
        "order_uid": order_uid,
        "order_number": order.order_number,
        "awb_count": len(awbs),
        "billable_count": billable_count,
        "billable_total": round(billable_cost, 2),
        "awbs": [
            {
                "id": awb.id,
                "tracking_number": awb.tracking_number,
                "courier_name": awb.courier_name,
                "awb_type": awb.awb_type or "outbound",
                "transport_cost": awb.transport_cost,
                "transport_cost_fara_tva": awb.transport_cost_fara_tva,
                "transport_cost_tva": awb.transport_cost_tva,
                "currency": awb.currency,
                "order_ref": awb.order_ref,
                "original_awb": awb.original_awb,
                "package_count": awb.package_count,
                "package_weight": awb.package_weight,
                "data_source": awb.data_source,
                "csv_status": awb.csv_status,
                "shipment_status": awb.shipment_status,
                "is_billable": is_billable_status(awb.csv_status),
                "created_at": to_bucharest_iso(awb.created_at),
            }
            for awb in awbs
        ]
    }


@router.patch("/{order_uid}/awb-count")
async def update_awb_count(
    order_uid: str,
    awb_count: int = Query(..., ge=1, le=10, description="Number of AWBs (1-10)"),
    db: AsyncSession = Depends(get_db)
):
    """Set the number of AWBs per order (1-10)."""
    result = await db.execute(select(Order).where(Order.uid == order_uid))
    order = result.scalar_one_or_none()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    order.awb_count = awb_count
    order.awb_count_manual = True
    await db.commit()
    
    return {
        "uid": order.uid,
        "awb_count": order.awb_count,
        "labels": [f"{i}/{awb_count}" for i in range(1, awb_count + 1)]
    }


@router.patch("/{order_uid}/shipping")
async def update_shipping_data(
    order_uid: str,
    package_count: Optional[int] = None,
    package_weight: Optional[float] = None,
    transport_cost: Optional[float] = None,
    db: AsyncSession = Depends(get_db)
):
    """Manually update shipping data for an order. Marks as manual to prevent CSV overwrite."""
    result = await db.execute(select(Order).where(Order.uid == order_uid))
    order = result.scalar_one_or_none()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    if package_count is not None:
        order.package_count = package_count
    if package_weight is not None:
        order.package_weight = package_weight
    if transport_cost is not None:
        order.transport_cost = transport_cost
    
    order.shipping_data_source = 'manual'
    order.shipping_data_manual = True
    await db.commit()
    
    return {
        "uid": order.uid,
        "package_count": order.package_count,
        "package_weight": order.package_weight,
        "transport_cost": order.transport_cost,
        "shipping_data_source": order.shipping_data_source,
    }


